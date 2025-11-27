import pandas as pd          # Manejo de tablas (CSV, Excel)
from openpyxl import load_workbook  # Para editar Excel
from openpyxl.styles import PatternFill, Font, Alignment  # Estilos
from openpyxl.utils import get_column_letter             # Ancho columnas
from openpyxl.worksheet.table import Table, TableStyleInfo  # Tablas Excel


# ===============================
# RUTAS DE ARCHIVOS (INPUT)
# ===============================

# Ruta del dataset de denuncias
ruta_csv = r"C:\era_data2\DATASET_Denuncias_Policiales_Enero 2018 a Octubre 2025.csv"

# Ruta del Excel con UBIGEO RENIEC que te mandó Karla
ruta_ubigeo = r"C:\era_data2\ubigeo.xlsx"   # <-- AJUSTA ESTA RUTA SI ES NECESARIO


# ===============================
# 1. CARGAR CSV (DENUNCIAS)
# ===============================

# Leer el CSV completo
df = pd.read_csv(ruta_csv, low_memory=False)

# Quitar espacios extra en los nombres de columnas
df.columns = df.columns.str.strip()

# Renombrar columnas para tener nombres cortos y consistentes
df = df.rename(columns={
    "UBIGEO_HECHO": "ubigeo",
    "ANIO": "año",
    "DPTO_HECHO_NEW": "departamento",
    "PROV_HECHO": "provincia",
    "DIST_HECHO": "distrito",
    "P_MODALIDADES": "modalidad",
    "cantidad": "cantidad"
})

# Asegurar formato estándar del UBIGEO en el CSV (6 dígitos)
df["ubigeo"] = df["ubigeo"].astype(str).str.zfill(6)


# ===============================
# 2. CARGAR UBIGEO RENIEC
# ===============================

ub = pd.read_excel(ruta_ubigeo)

# Asegurar que el ubigeo sea texto con 6 dígitos (ej: 10101 -> "010101")
ub["ubigeo"] = ub["ubigeo"].astype(str).str.zfill(6)

# Renombrar columnas del Excel de Karla
# Columnas del archivo: ubigeo, reg, prov, dist
ub = ub.rename(columns={
    "reg": "departamento_reniec",
    "prov": "provincia_reniec",
    "dist": "distrito_reniec"
})

# ===============================
# 3. UNIR CSV + UBIGEO RENIEC
# ===============================
# A partir de aquí, TODOS los nombres de dep/prov/dist vienen de RENIEC

df = df.merge(
    ub[["ubigeo", "departamento_reniec", "provincia_reniec", "distrito_reniec"]],
    on="ubigeo",
    how="left"
)

# Reemplazar por los nombres RENIEC (aunque ya existan columnas originales)
df["departamento"] = df["departamento_reniec"]
df["provincia"] = df["provincia_reniec"]
df["distrito"] = df["distrito_reniec"]

# Borrar columnas auxiliares RENIEC
df = df.drop(columns=["departamento_reniec", "provincia_reniec", "distrito_reniec"])


# ===============================
# FUNCIÓN PARA CREAR EXCEL FORMATEADO
# ===============================

def crear_excel_formateado(nombre_salida, dataframe):
    
    # Guardar Excel inicial
    dataframe.to_excel(nombre_salida, index=False, engine="openpyxl")
    wb = load_workbook(nombre_salida)
    ws = wb.active

    # Formato del encabezado
    header_fill = PatternFill("solid", fgColor="305496")   # Color azul
    header_font = Font(color="FFFFFF", bold=True)          # Letras blancas negrita
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Aplicar formato a cada columna del encabezado
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Crear tabla estilo Excel
    data_ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tabla = Table(displayName="Tabla_Datos", ref=data_ref)
    tabla.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tabla)

    # Autoajustar anchos de columna
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        max_len = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[col_letter].width = max_len + 2

    # Congelar primera fila
    ws.freeze_panes = "A2"

    wb.save(nombre_salida)
    print("✔ Archivo creado:", nombre_salida)


# ===============================
# 4. COLAPSO 1 (ANUAL–DISTRITAL–MODALIDAD)
# ===============================
# Suma por:
#  - año
#  - departamento (RENIEC)
#  - provincia (RENIEC)
#  - distrito (RENIEC)
#  - modalidad
#  - ubigeo (RENIEC)

colapso1 = (
    df.groupby(["año", "departamento", "provincia", "distrito", "modalidad", "ubigeo"])["cantidad"]
    .sum().reset_index()
)
crear_excel_formateado(r"C:\era_data2\colapso1.xlsx", colapso1)


# ===============================
# 5. COLAPSO 2 (TOTAL ANUAL)
# ===============================

colapso2 = df.groupby("año")["cantidad"].sum().reset_index()
crear_excel_formateado(r"C:\era_data2\colapso2.xlsx", colapso2)


# ===============================
# 6. COLAPSO 3 (AÑO × MODALIDAD)
# ===============================

colapso3 = df.groupby(["año", "modalidad"])["cantidad"].sum().reset_index()
crear_excel_formateado(r"C:\era_data2\colapso3.xlsx", colapso3)


# ===============================
# 6.5. COLAPSO 4 (AÑO × DISTRITO)
# ===============================

colapso4 = (
    df.groupby(["año", "departamento", "provincia", "distrito", "ubigeo"])["cantidad"]
    .sum().reset_index()
)
crear_excel_formateado(r"C:\era_data2\colapso4.xlsx", colapso4)


# ===============================
# 6.6. COLAPSO 5 (TOTAL POR DISTRITO)
# ===============================

colapso5 = (
    df.groupby(["departamento", "provincia", "distrito", "ubigeo"])["cantidad"]
    .sum().reset_index()
)
crear_excel_formateado(r"C:\era_data2\colapso5.xlsx", colapso5)


# ===============================
# 7. GANADOR CRIMINAL (DELITO PREDOMINANTE)
# ===============================

temp = df.groupby(
    ["ubigeo", "modalidad", "departamento", "provincia", "distrito"]
)["cantidad"].sum().reset_index()

# Elegimos el delito con mayor cantidad por distrito
ganador = temp.loc[temp.groupby("ubigeo")["cantidad"].idxmax()]
ganador = ganador.rename(columns={"modalidad": "modalidad_ganadora"})

crear_excel_formateado(r"C:\era_data2\ganador_criminal.xlsx", ganador)


# ===============================
# FINAL
# ===============================
print("\n✔ TODOS LOS COLAPSOS GENERADOS CORRECTAMENTE (USANDO UBIGEO RENIEC):")
print("   - colapso1.xlsx → año/provincia/distrito/modalidad (UBIGEO RENIEC)")
print("   - colapso2.xlsx → totales por año")
print("   - colapso3.xlsx → año × modalidad")
print("   - colapso4.xlsx → año × distrito (UBIGEO RENIEC)")
print("   - colapso5.xlsx → totales por distrito (UBIGEO RENIEC)")
print("   - ganador_criminal.xlsx → delito predominante por distrito (UBIGEO RENIEC)\n")
