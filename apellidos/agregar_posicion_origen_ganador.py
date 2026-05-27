"""
agregar_posicion_origen_ganador.py
===================================
Agrega tres variables a los CSVs de alcaldes_con_votos (2006-2022):

  1. posicion_cedula  – Posición del partido en la cédula de votación.
     Fuentes:
       • 2006-2014: Scrapping/data/base_final_con_posicion.csv
       • 2018-2022: Dato/BD/base_distrital_2018_2022_con_candidatos.csv
         (con fallback a base_final_con_posicion.csv)

  2. origen_apellido  – Clasificación lingüística del apellido paterno:
     QUECHUA, AYMARA, CASTELLANO, EXTRANJERO

  3. ganador  – 1 si es el candidato con más votos en su distrito (ubigeo)

Luego regenera los 5 Excel formateados.

Uso:
    python agregar_posicion_origen_ganador.py
"""

import pandas as pd
import unicodedata
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_DIR = os.path.dirname(BASE_DIR)

# ============================================================
# UTILIDADES
# ============================================================

def norm_text(s):
    """Normaliza texto: upper, sin tildes, sin ' - SIGLA' al final."""
    if pd.isna(s):
        return ""
    s = str(s).strip().upper()
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = re.sub(r"\s*-\s*[A-Z]{2,10}$", "", s)  # quitar " - FREPAP", " - PPC"
    return s.strip()


# ============================================================
# 1. MERGE DE POSICIÓN EN CÉDULA
# ============================================================

def load_position_sources():
    """Carga las fuentes de posición en cédula."""
    # Fuente A: base_final_con_posicion (todos los años, nivel distrito)
    bf = pd.read_csv(
        os.path.join(REPO_DIR, "Scrapping", "data", "base_final_con_posicion.csv"),
        dtype=str,
    )
    bf["ubigeo"] = bf["ubigeo"].str.zfill(6)
    bf["org_norm"] = bf["organizacion_politica"].apply(norm_text)
    bf["posicion"] = pd.to_numeric(bf["posicion"], errors="coerce")

    # Fuente B: base_distrital_2018_2022_con_candidatos (2018-2022, nivel candidato)
    bd_path = os.path.join(
        REPO_DIR, "Dato", "BD", "base_distrital_2018_2022_con_candidatos.csv"
    )
    bd = pd.read_csv(bd_path, dtype=str)
    # Normalizar nombre de columna año
    bd.columns = [c.replace("\ufeff", "").strip() for c in bd.columns]
    year_col = [c for c in bd.columns if "o" in c.lower() and "a" in c.lower()][0]
    bd = bd.rename(columns={year_col: "año"})
    bd["ubigeo"] = bd["ubigeo"].str.zfill(6)
    bd["org_norm"] = bd["organizacion_politica"].apply(norm_text)
    bd["orden_aparicion"] = pd.to_numeric(bd["orden_aparicion"], errors="coerce")

    return bf, bd


def merge_position(df, year, bf, bd):
    """Agrega posicion_cedula al dataframe de alcaldes."""
    year_str = str(year)
    df["org_norm"] = df["organizacion_politica"].apply(norm_text)

    if year <= 2014:
        # Usar solo base_final
        src = bf[bf["year"] == year_str][["ubigeo", "org_norm", "posicion"]].copy()
        src = src.drop_duplicates(subset=["ubigeo", "org_norm"])
        merged = df.merge(src, on=["ubigeo", "org_norm"], how="left")
        merged = merged.rename(columns={"posicion": "posicion_cedula"})
    else:
        # Usar base_distrital como primaria, base_final como fallback
        src_bd = (
            bd[bd["año"] == year_str][["ubigeo", "org_norm", "orden_aparicion"]]
            .copy()
            .drop_duplicates(subset=["ubigeo", "org_norm"])
        )
        src_bf = (
            bf[bf["year"] == year_str][["ubigeo", "org_norm", "posicion"]]
            .copy()
            .drop_duplicates(subset=["ubigeo", "org_norm"])
        )

        merged = df.merge(
            src_bd.rename(columns={"orden_aparicion": "pos_bd"}),
            on=["ubigeo", "org_norm"],
            how="left",
        )
        merged = merged.merge(
            src_bf.rename(columns={"posicion": "pos_bf"}),
            on=["ubigeo", "org_norm"],
            how="left",
        )
        merged["posicion_cedula"] = merged["pos_bd"].fillna(merged["pos_bf"])
        merged = merged.drop(columns=["pos_bd", "pos_bf"])

    merged = merged.drop(columns=["org_norm"])
    return merged


# ============================================================
# 2. CLASIFICACIÓN DE ORIGEN DE APELLIDO
# ============================================================

# --- Apellidos QUECHUA frecuentes (y raíces) ---
QUECHUA_EXACT = {
    "QUISPE", "HUAMAN", "CONDORI", "VILCA", "YUPANQUI", "SULLCA", "SULCA",
    "HUANCA", "HUALLPA", "HUILLCA", "CHALLCO", "PILLCO", "PUMACAHUA",
    "CCAHUANA", "CCAMA", "CCOPA", "CCARI", "CCASA", "CCALLA", "CCALLO",
    "CCAHUA", "CCASANI", "CCORIMANYA", "CCOLQUE", "CCORAHUA",
    "TTITO", "TTICA", "TTURO",
    "QQUELLON", "QQUESHUAYLLO",
    "PHOCCO", "PHAJSI",
    "HUAMANI", "HUAMANTICA", "HUAMANLAZO", "HUAMANTUPA",
    "HUAYTA", "HUAYHUA", "HUAYTALLA",
    "ATAUCHI", "ATAUCURI", "ATAUCUSI",
    "CUSIHUAMAN", "CUSIYUPANQUI",
    "INCA", "INCACUTIPA", "INCAHUANACO",
    "SINCHI", "SONCCO",
    "TUPA", "TUPAC", "TUPAYACHI",
    "PAUCAR", "PAUCARA", "PAUCCAR",
    "TICLLA", "TICLLASUCA",
    "PALOMINO", "TAIPE", "TAYPE",
    "CCOLQQUE", "CCOLQUEHUANCA",
    "HUARAYA", "HUARACHI", "HUARACALLO",
    "CHUQUIMIA", "CHUQUITAPA", "CHUQUIHUANCA",
    "ÑAHUI", "ÑAUPA", "ÑAHUINLLA",
    "PUMACCAHUA", "PUMA", "PUMAHUALLCA",
    "PILLACA", "PILLACA", "PILLPE",
    "QQUENTA", "QQENAYA",
    "CHILLITUPA", "CHILLIHUANI",
    "SACCSA", "SACCATOMA",
    "AUCCA", "AUCCAPUMA", "AUCCAPIÑA",
    "HUISA", "HUISACAYNA",
    "RAMOS", # ambiguo pero frecuente en zona quechua
    "TICONA",  # compartido Q/A
    "HANCCO", "HALLASI",
    "QUISBERT",  # altiplano
    "CHOQUE",  # compartido Q/A
    "CHOQUEHUANCA",
    "COLQUE", "CCOLQUE", "COLQUEHUANCA",
    "HUANCOLLO",
    "LLALLACACHI",
    "MAYTA",
    "MACHACA",
    "PACOMPIA",
    "QUILLE", "QUILLA",
    "TURPO", "TURPAUD",
    "USCA",
    "WANKA",
    "YUCRA",
}

QUECHUA_ROOTS = [
    "HUAMAN", "QUISP", "VILCA", "CONDOR", "CHALLC", "PILLC", "PUMAC",
    "HUILLC", "CHUQUI", "HUAYLLA", "CCAHUA", "CCALL", "CCOPA",
    "HUARAC", "YUPANQ", "CUSIH", "PAUCC", "TTICL", "SONCCO",
    "ATAUCU", "ÑAHUI", "ÑAUP", "SACCS", "AUCCA", "QQUENT",
    "CCORI", "TTUPA", "CCAMA", "PUMAHUA", "SULLCA", "HUANCA",
    "HUAYTA", "HUALLP",
]

# --- Apellidos AYMARA frecuentes ---
AYMARA_EXACT = {
    "MAMANI", "APAZA", "TICONA", "CHOQUE", "ALANOCA", "CATACORA",
    "COILA", "MAQUERA", "LUPACA", "CUTIPA", "PACOTICONA", "CHAMBILLA",
    "LARICO", "LLANQUE", "LLANOS", "CHARAJA", "CALIZAYA",
    "CAHUANA", "COPARI", "CHURA", "CHINO", "CHURATA",
    "CONDEMAYTA", "CALCINA", "CAUNA", "CARI", "CANAZA",
    "CHIPANA", "COAQUIRA", "COAYLA", "COLLANQUI",
    "ESCOBAR",  # no, quitar
    "HALLPA", "HUARACHA", "HUAYNACHO", "INQUILLA",
    "LAQUI", "LIMACHI", "LUQUE", "MARCA",
    "NINA", "PARICAHUA", "PILCO",
    "PONCE",  # ambiguo
    "QUEA", "QUELCA", "QUENTA",
    "ROQUE", "SACARI", "SUCA",
    "TACO", "TICAHUANCA", "TINTAYA",
    "UCHASARA", "VILCANQUI",
    "YANAPA", "YUJRA", "ZAPANA",
    "SULLA", "SUCAPUCA",
    "AROHUANCA", "AROAPAZA",
    "ALIAGA",  # ambiguo, quitar
    "ACHATA", "ACHAHUI",
    "ADCO", "ADUVIRI",
    "AGUILAR",  # no, castellano
    "ARPASI",
    "CALLO", "CALLATA",
    "CCALLATA",
    "HUAQUISTO",
    "INQUILTUPA",
    "JALANOCA",
    "NINAJA", "NINARAQUI",
    "PACHO", "PACHARI",
    "PHALA",
    "SURCO", "SUPO",
    "TIPO", "TOLA",
    "UCHIRI",
}

# Quitar ambiguos/españoles del set Aymara
AYMARA_EXACT -= {"ESCOBAR", "ALIAGA", "AGUILAR"}

AYMARA_ROOTS = [
    "MAMANI", "APAZA", "ALANOC", "CATACO", "MAQUE", "LUPAC",
    "CUTIP", "CHAMBI", "LARIC", "LLANQU", "CHARAJ", "CALIZA",
    "COPARI", "CHURAT", "CALCIN", "CANAZ", "CHIPAN", "COAQUI",
    "COLLAN", "INQUIL", "LIMACH", "PARICA", "QUELCA",
    "SACARI", "SUCAPU", "AROHUA", "ADUVIR", "CALLAT",
    "HUAQUI", "JALANO", "NINARA", "PACHAR",
]

# --- Apellidos EXTRANJERO (no hispanos) ---
EXTRANJERO_EXACT = {
    # Japoneses
    "FUJIMORI", "HIGUCHI", "HIGA", "MIYASHIRO", "NAKAMURA", "YAMAMOTO",
    "TANAKA", "SHIMABUKURO", "IKEDA", "KOBAYASHI", "TAKAHASHI",
    "MATSUDA", "YOSHIDA", "WATANABE", "SUZUKI", "KIMURA", "SAITO",
    "MORIMOTO", "OKADA", "TSUBOYAMA", "NISHIYAMA",
    # Chinos
    "WONG", "CHANG", "CHEN", "CHENG", "CHUNG", "CHOY", "CHIU",
    "FUNG", "HUANG", "KWONG", "LAM", "LAU", "LEE", "LI", "LIN",
    "LIU", "LUNG", "SIU", "TAM", "TANG", "TONG", "WU", "YANG",
    "YEN", "YIP",
    # Italianos
    "FERRARI", "FERRARO", "BERTOLOTTO", "MUSSOLINI", "PAOLETTI",
    "BIANCHI", "MORELLI", "ZANETTI", "ROSSI",
    "BERETTA", "COLOMBO", "CONTI", "COSTA", "DAMIANI",
    "DELL", "GALLI", "GIORDANO", "LOMBARDI", "MANCINI",
    "MARCHETTI", "MARTINI", "MAZZA", "NEGRO", "PAGANI",
    "PELLEGRINI", "RICCI", "RINALDI", "ROMANO", "RUSSO",
    "SANTINI", "SERRA", "VITALE",
    # Alemanes
    "SCHMIDT", "SCHNEIDER", "MULLER", "WEBER", "MEYER", "FISCHER",
    "HOFFMANN", "BAUER", "SCHULZ", "KOCH",
    # Ingleses
    "SMITH", "JOHNSON", "BROWN", "WILLIAMS", "JONES", "MILLER",
    "TAYLOR", "ANDERSON", "THOMAS", "JACKSON", "WHITE",
    # Portugueses/Brasileños
    "PEREIRA", "FERREIRA", "ALMEIDA", "OLIVEIRA", "RIBEIRO",
    "CARDOSO", "CARVALHO", "GOMES", "PINTO", "TEIXEIRA",
    "FREITAS", "MACEDO", "MOREIRA", "FONSECA", "VIEIRA",
    # Franceses
    "DUPONT", "DURAND", "LAURENT", "LEFEVRE", "MARTIN",
    "BERNARD", "PETIT", "ROBERT",
    # Árabes/Turcos
    "SALOMON", "ABUSADA", "MAJLUF", "MUFARECH",
    # Otros
    "BELMONT", "KUCZYNSKI",
}

EXTRANJERO_SUFFIXES = [
    "MOTO", "SHIRO", "MURA", "GUCHI", "KURO", "YAMA", "GAWA",
    "ISHI", "ZAKI", "ZAWA", "NAGA", "UCHI",  # japoneses
    "ETTI", "OTTI", "ELLI", "ACCI", "UCCI", "INI",  # italianos
    "FELD", "BERG", "BURG", "MANN", "STEIN",  # alemanes
    "EIRA", "EIRA",  # portugueses
]

# --- Sufijos castellanos ---
CASTELLANO_SUFFIXES = [
    "EZ", "AZ", "OZ", "IZ",   # patronímicos: López, Díaz, Muñoz
    "ON", "ION",                # Morón, Girón
    "ERO", "ERA",               # Romero, Herrera
    "ADO", "IDA", "IDO",       # Delgado, Partida
    "ILLO", "ILLA",            # Carrillo, Padilla
    "ALES", "ARES",            # Morales, Linares
    "ERE", "ERO",
    "ENTE",                    # Valiente, Puente
    "ANDA", "ONDO",            # Miranda, Redondo
    "UEL",                     # Miguel, Manuel
    "TION",
    "ANZA", "ENZA",            # Carranza, Atienza
    "ASCO", "ESCO",            # Velasco, Tedesco
]

CASTELLANO_COMMON = {
    "GARCIA", "RODRIGUEZ", "MARTINEZ", "LOPEZ", "GONZALEZ", "HERNANDEZ",
    "PEREZ", "SANCHEZ", "RAMIREZ", "TORRES", "FLORES", "RIVERA",
    "GOMEZ", "DIAZ", "REYES", "MORALES", "JIMENEZ", "RUIZ",
    "ALVAREZ", "GUTIERREZ", "ORTIZ", "MENDOZA", "CASTILLO", "ROMERO",
    "CRUZ", "VASQUEZ", "FERNANDEZ", "VARGAS", "MEDINA", "HERRERA",
    "CASTRO", "SILVA", "CHAVEZ", "ROJAS", "VEGA", "ESPINOZA",
    "RAMOS", "CORDOVA", "CAMPOS", "DELGADO", "LEON", "AGUILAR",
    "SALAZAR", "PAREDES", "SOTO", "RIOS", "ACOSTA", "VERA",
    "VALDEZ", "VALVERDE", "CARBAJAL", "BENITES", "BENITEZ",
    "CARRASCO", "CAMACHO", "CERVANTES", "CISNEROS", "CONTRERAS",
    "CORONEL", "DAVILA", "DOMINGUEZ", "DURAN", "ENRIQUEZ",
    "ESCOBAR", "ESQUIVEL", "FIGUEROA", "GALLEGOS", "GUERRERO",
    "HIDALGO", "HUERTA", "IBARRA", "LARA", "LUNA", "MALDONADO",
    "MARIN", "MARQUEZ", "MEZA", "MIRANDA", "MONTOYA", "NAVARRO",
    "NORIEGA", "NUNES", "NUNEZ", "OCHOA", "OLIVA", "OSORIO",
    "PADILLA", "PALACIOS", "PENA", "PINEDA", "POLO", "PORTILLO",
    "QUIROZ", "RETAMOZO", "ROSALES", "SALCEDO", "SANTANA",
    "SEGURA", "SERRANO", "SOLIS", "SUAREZ", "TAPIA", "TEJADA",
    "TRUJILLO", "URBINA", "VALENZUELA", "VALLEJO", "VELASQUEZ",
    "VIDAL", "VILLANUEVA", "VILLEGAS", "ZARATE", "ZAVALA", "ZUNIGA",
    "ABANTO", "ADRIANZEN", "ALARCON", "ALCANTARA", "ALIAGA", "ALVA",
    "ARCE", "AREVALO", "ARIAS", "ARROYO", "AVALOS", "AVILA",
    "BARRANTES", "BARRERA", "BAUTISTA", "BECERRA", "BERMUDEZ",
    "BRAVO", "BRICE", "CABRERA", "CALDERON", "CARRION",
    "CASAS", "CASTANEDA", "CERRON", "CHAVEZ", "CORDERO",
    "CUEVA", "CUEVAS", "DE LA CRUZ", "DIEZ", "ESTRADA",
    "FALCON", "GALARZA", "GALLARDO", "GAMBOA", "GARAY",
    "GIL", "GUEVARA", "HEREDIA", "HOYOS", "IZQUIERDO",
    "JARA", "JULCA", "JURADO", "LAMAS", "LEDESMA",
    "LINARES", "LLANOS", "LLERENA", "LOZANO", "MACHADO",
    "MANTILLA", "MATOS", "MEJIA", "MELENDEZ", "MOLINA",
    "MONTES", "MONTERO", "MORON", "MUÑOZ", "MUNOZ",
    "OBREGON", "OJEDA", "OLIVOS", "ORELLANA", "ORIHUELA",
    "ORTEGA", "OSSA", "OTERO", "PARDO", "PASTOR",
    "PERALTA", "PRADO", "PUENTE", "QUIJANO", "QUINTANA",
    "RABANAL", "RENGIFO", "REATEGUI", "RICALDI", "RIOJA",
    "RIVA", "ROCA", "ROMAN", "RUBIO", "SAENZ",
    "SALAS", "SALVADOR", "SARMIENTO", "SEMINARIO", "SOLANO",
    "SORIANO", "TAMAYO", "TELLO", "TOLEDO", "TORREJON",
    "UGARTE", "ULLOA", "URQUIZO", "VALENCIA", "VALERA",
    "VARELA", "VASQUEZ", "VENTURA", "VERGARA", "VICTORIA",
    "VILCHEZ", "VILLALBA", "YBANEZ", "ZAVALETA", "ZEVALLOS",
}


def clasificar_origen(apellido):
    """Clasifica el origen lingüístico de un apellido peruano."""
    if pd.isna(apellido) or str(apellido).strip() == "":
        return "CASTELLANO"

    ap = norm_text(apellido)
    if not ap:
        return "CASTELLANO"

    # 1. Extranjero exacto
    if ap in EXTRANJERO_EXACT:
        return "EXTRANJERO"

    # 2. Sufijos extranjeros
    for suf in EXTRANJERO_SUFFIXES:
        if len(ap) > len(suf) + 2 and ap.endswith(suf):
            # Evitar falsos positivos con castellano
            if not any(ap.endswith(cs) for cs in CASTELLANO_SUFFIXES):
                return "EXTRANJERO"

    # 3. Quechua: doble consonante al inicio (CC, TT, QQ, PP, LL)
    if re.match(r"^(CC|TT|QQ|PP)[A-Z]", ap):
        return "QUECHUA"

    # 4. Quechua: PH al inicio
    if ap.startswith("PH"):
        return "QUECHUA"

    # 5. Quechua exacto
    if ap in QUECHUA_EXACT:
        return "QUECHUA"

    # 6. Raíces quechua
    for root in QUECHUA_ROOTS:
        if root in ap and len(ap) >= len(root):
            return "QUECHUA"

    # 7. Aymara exacto
    if ap in AYMARA_EXACT:
        return "AYMARA"

    # 8. Raíces aymara
    for root in AYMARA_ROOTS:
        if root in ap and len(ap) >= len(root):
            return "AYMARA"

    # 9. Castellano exacto
    if ap in CASTELLANO_COMMON:
        return "CASTELLANO"

    # 10. Sufijos castellanos
    for suf in CASTELLANO_SUFFIXES:
        if ap.endswith(suf) and len(ap) > len(suf) + 1:
            return "CASTELLANO"

    # 11. Default
    return "CASTELLANO"


# ============================================================
# 3. VARIABLE GANADOR
# ============================================================

def agregar_ganador(df):
    """Marca ganador=1 al candidato con más votos en cada ubigeo."""
    df = df.copy()
    df["_votos_num"] = pd.to_numeric(df["total_votos"], errors="coerce")
    df["ganador"] = 0

    for ubigeo, grp in df.groupby("ubigeo"):
        if grp["_votos_num"].notna().any():
            max_idx = grp["_votos_num"].idxmax()
            df.loc[max_idx, "ganador"] = 1

    df = df.drop(columns=["_votos_num"])
    return df


# ============================================================
# PROCESAMIENTO PRINCIPAL
# ============================================================

print("Cargando fuentes de posición en cédula...")
bf, bd = load_position_sources()

YEARS = [2006, 2010, 2014, 2018, 2022]

for year in YEARS:
    csv_path = os.path.join(BASE_DIR, "alcaldes_con_votos", f"alcaldes_con_votos_{year}.csv")
    print(f"\n{'='*60}")
    print(f"  AÑO {year}")
    print(f"{'='*60}")

    df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
    df.columns = [c.strip() for c in df.columns]

    # Quitar columnas previas si existen (re-run safe)
    for col in ["posicion_cedula", "origen_apellido", "ganador"]:
        if col in df.columns:
            df = df.drop(columns=[col])

    n_total = len(df)

    # --- Posición en cédula ---
    df = merge_position(df, year, bf, bd)
    n_pos = df["posicion_cedula"].notna().sum()
    print(f"  Posición: {n_pos}/{n_total} ({n_pos/n_total*100:.1f}%)")

    # --- Origen de apellido ---
    df["origen_apellido"] = df["apellido_paterno"].apply(clasificar_origen)
    dist = df["origen_apellido"].value_counts()
    print(f"  Origen: {dist.to_dict()}")

    # --- Ganador ---
    df = agregar_ganador(df)
    n_gan = df["ganador"].sum()
    n_dist = df["ubigeo"].nunique()
    print(f"  Ganadores: {n_gan} (distritos: {n_dist})")

    # --- Reordenar columnas ---
    col_order = [
        "ubigeo", "departamento", "provincia", "distrito",
        "organizacion_politica", "cargo",
        "nombres", "apellido_paterno", "apellido_materno",
        "sexo", "total_votos", "posicion_cedula", "origen_apellido", "ganador",
    ]
    # Incluir columnas extras que puedan existir (ej: dni)
    extras = [c for c in df.columns if c not in col_order]
    col_order = [c for c in col_order if c in df.columns] + extras
    df = df[col_order]

    # --- Guardar ---
    df.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"  Guardado: {csv_path}")

    # Mostrar ejemplos
    sample = df[df["posicion_cedula"].notna()].head(3)
    for _, r in sample.iterrows():
        print(f"    {r['nombres']:<20} {r['apellido_paterno']:<15} pos={r['posicion_cedula']}"
              f"  origen={r['origen_apellido']:<12} ganador={r['ganador']}")


# ============================================================
# REGENERAR EXCEL
# ============================================================

print(f"\n{'='*60}")
print("Regenerando 5 archivos Excel...")
print(f"{'='*60}")

CSV_DIR = os.path.join(BASE_DIR, "alcaldes_con_votos")

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
EVEN_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
CENTER_COLS = {"ubigeo", "sexo", "total_votos", "dni", "posicion_cedula", "ganador"}

for year in YEARS:
    csv_path = os.path.join(CSV_DIR, f"alcaldes_con_votos_{year}.csv")
    xlsx_path = os.path.join(BASE_DIR, f"alcaldes_{year}.xlsx")

    df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
    df.columns = [c.strip() for c in df.columns]

    df.to_excel(xlsx_path, index=False, sheet_name=f"Alcaldes {year}", engine="openpyxl")

    wb = load_workbook(xlsx_path)
    ws = wb.active
    num_cols = ws.max_column
    num_rows = ws.max_row
    headers = [ws.cell(row=1, column=c).value for c in range(1, num_cols + 1)]

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx in range(2, num_rows + 1):
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.fill = fill
            col_name = headers[col_idx - 1]
            if col_name and col_name.lower() in CENTER_COLS:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

    for col_idx in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(num_rows + 1, 1000)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        adjusted = min(max_len + 3, 45)
        ws.column_dimensions[col_letter].width = max(adjusted, 8)

    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows}"
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    print(f"  {year}: {xlsx_path} ({num_rows - 1} filas)")

print("\n¡Listo!")
