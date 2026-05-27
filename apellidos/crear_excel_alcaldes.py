"""
crear_excel_alcaldes.py
=======================
Genera 5 archivos Excel formateados (.xlsx) a partir de los CSVs
de alcaldes_con_votos/ para cada año electoral (2006–2022).

Formato aplicado:
- Autofiltro en todas las columnas
- Encabezados en negrita, fondo azul oscuro (#1F3864), texto blanco
- Columnas con ancho ajustado al contenido
- Bordes en todas las celdas
- Ubigeo, sexo, dni y total_votos centrados
- Filas alternas con fondo azul claro para facilitar lectura
- Primera fila congelada (freeze panes)

Uso:
    python crear_excel_alcaldes.py

Entrada:  apellidos/alcaldes_con_votos/alcaldes_con_votos_YYYY.csv
Salida:   apellidos/alcaldes_YYYY.xlsx
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CSV_DIR = os.path.join(BASE_DIR, "alcaldes_con_votos")

YEARS = [2006, 2010, 2014, 2018, 2022]

# Estilo
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

EVEN_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")

# Columnas que deben ir centradas
CENTER_COLS = {"ubigeo", "sexo", "total_votos", "dni"}

for year in YEARS:
    csv_path = os.path.join(CSV_DIR, f"alcaldes_con_votos_{year}.csv")
    xlsx_path = os.path.join(BASE_DIR, f"alcaldes_{year}.xlsx")

    print(f"Procesando {year}...")
    df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)

    # Limpiar nombres de columnas
    df.columns = [c.strip() for c in df.columns]

    # Escribir a Excel sin formato primero
    df.to_excel(xlsx_path, index=False, sheet_name=f"Alcaldes {year}", engine="openpyxl")

    # Abrir y formatear
    wb = load_workbook(xlsx_path)
    ws = wb.active

    num_cols = ws.max_column
    num_rows = ws.max_row
    headers = [ws.cell(row=1, column=c).value for c in range(1, num_cols + 1)]

    # Formato de encabezados
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Formato de filas de datos
    for row_idx in range(2, num_rows + 1):
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.fill = fill
            col_name = headers[col_idx - 1]
            if col_name and col_name.lower() in CENTER_COLS:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

    # Ajustar ancho de columnas
    for col_idx in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(num_rows + 1, 1000)):  # sample first 1000 rows
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        # Cap width and add padding
        adjusted = min(max_len + 3, 45)
        ws.column_dimensions[col_letter].width = max(adjusted, 8)

    # Autofiltro
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows}"

    # Congelar primera fila
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    print(f"  -> {xlsx_path} ({num_rows - 1} filas)")

print("\nListo! 5 archivos Excel creados.")
