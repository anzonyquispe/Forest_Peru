"""
fix_merge_y_apellidos.py
========================
Corrige los mapeos de partidos JNE->ONPE, rehace el merge de alcaldes
con votos para los 5 años electorales, separa nombres para 2006-2014,
y regenera los 5 archivos Excel formateados.

Cambios respecto al merge original:
- ORG_MAP_2018: elimina 2 mapeos rotos, agrega 13 nuevos
- ORG_MAP_2022: agrega 1 mapeo nuevo (HATARIY APURIMAC)
- Separación de nombres para 2006/2010/2014 desde el nombre completo

Uso:
    python fix_merge_y_apellidos.py
"""

import pandas as pd
import unicodedata
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# --Paths --------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(BASE_DIR)
SCRAPPING = os.path.join(REPO, "Scrapping")
PARTIDOS = os.path.join(SCRAPPING, "partidos_regionales")
CANDIDATOS_SRC = os.path.join(SCRAPPING, "partidos_regionales", "candidatos", "output")
CANDIDATOS_DEST = os.path.join(BASE_DIR, "candidatos_completos")
ALCALDES_DIR = os.path.join(BASE_DIR, "alcaldes_con_votos")

# --Party name mappings ------------------------------------------------
# 2018: 2 correctos + 13 nuevos (2 rotos eliminados)
ORG_MAP_2018 = {
    # Existentes correctos
    "PODEMOS PERU": "PODEMOS POR EL PROGRESO DEL PERU",
    "FRENTE AMPLIO POR JUSTICIA, VIDA Y LIBERTAD": "FRENTE AMPLIO",
    # Nuevos (descubiertos por análisis de co-ocurrencia por ubigeo)
    "VICTORIA NACIONAL": "RESTAURACION NACIONAL",
    "RENACIMIENTO UNIDO NACIONAL": "SIEMPRE UNIDOS",
    "PARTIDO POLITICO NACIONAL PERU LIBRE": "PERU LIBERTARIO",
    "PARTIDO POLITICO CONTIGO": "PERUANOS POR EL KAMBIO",
    "RENOVACION POPULAR": "SOLIDARIDAD NACIONAL",
    "MOVIMIENTO REGIONAL WARI LLAQTA": "TECNOLOGIA DE PUNTA PARA AYACUCHO",
    "MOVIMIENTO REGIONAL VIVA TACNA": "ACCION POR LA UNIDAD TACNA",
    "MOVIMIENTO REGIONAL MAS CALLAO": "FUERZA CHALACA",
    "MOVIMIENTO REGIONAL CONSTRUYENDO": "PRIMERO LAMBAYEQUE",
    "SI SE PUEDE": "FRENTE UNITARIO POPULAR",
    "MR FUERZA POR MADRE DE DIOS - MR FMDD": "FUERZA POR MADRE DE DIOS",
}

ORG_MAP_2022 = {
    "HATARIY APURIMAC": "MOVIMIENTO REGIONAL HATARIY APURIMAC",
}

# --Utility functions --------------------------------------------------

def strip_accents(text):
    if not isinstance(text, str):
        return str(text) if text is not None else ""
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(ch)
    )


def normalize_org(name):
    return strip_accents(str(name)).upper().strip()


def read_csv_auto_encoding(path, **kwargs):
    """Read CSV trying utf-8-sig, utf-8, then latin-1."""
    for enc in ["utf-8-sig", "utf-8", "latin-1"]:
        try:
            return pd.read_csv(path, encoding=enc, **kwargs)
        except UnicodeDecodeError:
            continue
    raise UnicodeDecodeError(f"Could not decode {path}")


def split_name(full_name):
    """Split a full name string into nombres, apellido_paterno, apellido_materno.

    Rules:
    - 3+ words: last 2 -> apellido_paterno + apellido_materno, rest -> nombres
    - 2 words: first -> nombres, second -> apellido_paterno
    - 1 word: -> nombres only
    """
    parts = str(full_name).strip().split() if pd.notna(full_name) else []
    if len(parts) >= 3:
        return {
            "nombres": " ".join(parts[:-2]),
            "apellido_paterno": parts[-2],
            "apellido_materno": parts[-1],
        }
    elif len(parts) == 2:
        return {
            "nombres": parts[0],
            "apellido_paterno": parts[1],
            "apellido_materno": "",
        }
    elif len(parts) == 1:
        return {
            "nombres": parts[0],
            "apellido_paterno": "",
            "apellido_materno": "",
        }
    else:
        return {"nombres": "", "apellido_paterno": "", "apellido_materno": ""}


def get_votos_path(year):
    if year <= 2014:
        return os.path.join(PARTIDOS, f"distrital_{year}.csv")
    elif year == 2018:
        return os.path.join(SCRAPPING, "resultados_onpe_erm2018_por_distrito_distrital.csv")
    else:
        return os.path.join(SCRAPPING, "resultados_onpe_erm2022_por_distrito_distrital.csv")


def get_org_map(year):
    if year == 2018:
        return ORG_MAP_2018
    elif year == 2022:
        return ORG_MAP_2022
    return {}


# --Columns for output -------------------------------------------------
CANDIDATOS_COLS_NO_DNI = [
    "ubigeo", "departamento", "provincia", "distrito",
    "organizacion_politica", "cargo",
    "nombres", "apellido_paterno", "apellido_materno", "sexo",
]
CANDIDATOS_COLS_DNI = CANDIDATOS_COLS_NO_DNI + ["dni"]

ALCALDES_COLS_NO_DNI = CANDIDATOS_COLS_NO_DNI + ["total_votos"]
ALCALDES_COLS_DNI = CANDIDATOS_COLS_DNI + ["total_votos"]


# --Excel formatting (same as crear_excel_alcaldes.py) -----------------
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
EVEN_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
CENTER_COLS = {"ubigeo", "sexo", "total_votos", "dni"}


def format_excel(xlsx_path, year):
    """Apply formatting to an Excel file (same as crear_excel_alcaldes.py)."""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    num_cols = ws.max_column
    num_rows = ws.max_row
    headers = [ws.cell(row=1, column=c).value for c in range(1, num_cols + 1)]

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx in range(2, num_rows + 1):
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.fill = fill
            col_name = headers[col_idx - 1]
            if col_name and col_name.lower() in CENTER_COLS:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

    for col_idx in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(num_rows + 1, 1000)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        adjusted = min(max_len + 3, 45)
        ws.column_dimensions[col_letter].width = max(adjusted, 8)

    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows}"
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)


# =======================================================================
# MAIN
# =======================================================================

def main():
    years = [2006, 2010, 2014, 2018, 2022]

    # Collect before/after stats
    before_stats = {}
    for year in years:
        path = os.path.join(ALCALDES_DIR, f"alcaldes_con_votos_{year}.csv")
        if os.path.exists(path):
            df = pd.read_csv(path, dtype=str)
            n = len(df)
            matched = df["total_votos"].notna() & (df["total_votos"].astype(str).str.strip() != "")
            before_stats[year] = (n, matched.sum())

    print("=" * 70)
    print("fix_merge_y_apellidos.py - Corregir mapeos, rehacer merge, Excel")
    print("=" * 70)

    # --Step 1: Split names for 2006/2010/2014 candidatos_completos ----
    print("\n--Paso 1: Separar nombres en candidatos_completos (2006-2014) --")
    for year in [2006, 2010, 2014]:
        src_path = os.path.join(CANDIDATOS_SRC, f"candidatos_{year}.csv")
        dest_path = os.path.join(CANDIDATOS_DEST, f"candidatos_{year}.csv")

        df = read_csv_auto_encoding(src_path, dtype=str)
        print(f"\n  {year}: {len(df)} candidatos leidos de scraping output")

        # Split full names from 'nombres' column (apellido_paterno/materno are empty)
        splits = df["nombres"].apply(split_name).apply(pd.Series)
        df["nombres"] = splits["nombres"]
        df["apellido_paterno"] = splits["apellido_paterno"]
        df["apellido_materno"] = splits["apellido_materno"]

        # Keep only the standard columns (no dni for 2006-2014, no sexo_fuente)
        out_cols = [c for c in CANDIDATOS_COLS_NO_DNI if c in df.columns]
        df[out_cols].to_csv(dest_path, index=False, encoding="utf-8-sig")
        print(f"    -> {dest_path}")
        # Show sample
        sample = df.iloc[0]
        print(f"    Ejemplo: '{sample['nombres']}' | '{sample['apellido_paterno']}' | '{sample['apellido_materno']}'")

    # --Step 2: Merge alcaldes con votos for all 5 years ---------------
    print("\n--Paso 2: Merge alcaldes con votos ONPE --")
    after_stats = {}

    for year in years:
        print(f"\n  {year}:")

        # Read candidatos source
        if year <= 2014:
            # Use the scraping output (full names) and split
            cand_path = os.path.join(CANDIDATOS_SRC, f"candidatos_{year}.csv")
            df_cand = read_csv_auto_encoding(cand_path, dtype=str)
            # Split names
            splits = df_cand["nombres"].apply(split_name).apply(pd.Series)
            df_cand["nombres"] = splits["nombres"]
            df_cand["apellido_paterno"] = splits["apellido_paterno"]
            df_cand["apellido_materno"] = splits["apellido_materno"]
        else:
            # 2018/2022: names already split in scraping output
            cand_path = os.path.join(CANDIDATOS_SRC, f"candidatos_{year}.csv")
            df_cand = read_csv_auto_encoding(cand_path, dtype=str)

        # Filter to ALCALDE DISTRITAL
        alcaldes = df_cand[
            df_cand["cargo"].astype(str).str.upper().str.contains("ALCALDE DISTRITAL", na=False)
        ].copy()
        print(f"    Alcaldes: {len(alcaldes)}")

        # Read ONPE votes
        votos_path = get_votos_path(year)
        df_votos = read_csv_auto_encoding(votos_path, dtype=str)
        df_votos["ubigeo"] = df_votos["ubigeo"].astype(str).str.zfill(6)

        # Filter out summary rows
        df_votos = df_votos[
            ~df_votos["organizacion_politica"].astype(str).str.contains(
                "TOTAL|VOTOS EN BLANCO|VOTOS NULOS|VOTOS IMPUGNADOS",
                na=False, case=False,
            )
        ]

        # Normalize org names for matching
        alcaldes["org_norm"] = alcaldes["organizacion_politica"].apply(normalize_org)
        df_votos["org_norm"] = df_votos["organizacion_politica"].apply(normalize_org)

        # Apply party name mapping
        org_map = get_org_map(year)
        if org_map:
            mapped_count = 0
            for old, new in org_map.items():
                mask = alcaldes["org_norm"] == normalize_org(old)
                mapped_count += mask.sum()
                alcaldes.loc[mask, "org_norm"] = normalize_org(new)
            print(f"    Mapeos aplicados: {len(org_map)} reglas, {mapped_count} filas afectadas")

        # Convert total_votos to numeric in df_votos
        df_votos["total_votos"] = pd.to_numeric(df_votos["total_votos"], errors="coerce")

        # Merge
        merged = alcaldes.merge(
            df_votos[["ubigeo", "org_norm", "total_votos"]],
            on=["ubigeo", "org_norm"],
            how="left",
        )

        # Drop helper column
        merged = merged.drop(columns=["org_norm"])

        # Select output columns
        if year >= 2018:
            out_cols = [c for c in ALCALDES_COLS_DNI if c in merged.columns]
        else:
            out_cols = [c for c in ALCALDES_COLS_NO_DNI if c in merged.columns]

        out_path = os.path.join(ALCALDES_DIR, f"alcaldes_con_votos_{year}.csv")
        merged[out_cols].to_csv(out_path, index=False, encoding="utf-8-sig")

        n = len(merged)
        n_matched = merged["total_votos"].notna().sum()
        pct = n_matched / n * 100 if n > 0 else 0
        after_stats[year] = (n, n_matched)
        print(f"    Match: {n_matched}/{n} ({pct:.1f}%)")
        print(f"    -> {out_path}")

    # --Step 3: Split names in alcaldes_con_votos for 2006-2014 --------
    # Already done during merge (step 2), no separate step needed

    # --Step 4: Regenerate Excel files ---------------------------------
    print("\n--Paso 3: Regenerar archivos Excel --")
    for year in years:
        csv_path = os.path.join(ALCALDES_DIR, f"alcaldes_con_votos_{year}.csv")
        xlsx_path = os.path.join(BASE_DIR, f"alcaldes_{year}.xlsx")

        df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        df.to_excel(xlsx_path, index=False, sheet_name=f"Alcaldes {year}", engine="openpyxl")
        format_excel(xlsx_path, year)
        print(f"  {year}: {xlsx_path} ({len(df)} filas)")

    # --Step 5: Summary ------------------------------------------------
    print("\n" + "=" * 70)
    print("COMPARACION ANTES / DESPUES")
    print("=" * 70)
    print(f"\n{'Year':<6} {'Antes':>16} {'Despues':>16} {'D Matches':>10}")
    print("-" * 52)
    for year in years:
        b_n, b_m = before_stats.get(year, (0, 0))
        a_n, a_m = after_stats.get(year, (0, 0))
        b_pct = f"{b_m}/{b_n} ({b_m/b_n*100:.1f}%)" if b_n > 0 else "N/A"
        a_pct = f"{a_m}/{a_n} ({a_m/a_n*100:.1f}%)" if a_n > 0 else "N/A"
        delta = a_m - b_m
        print(f"{year:<6} {b_pct:>16} {a_pct:>16} {'+' if delta >= 0 else ''}{delta:>9}")

    # --Step 6: Sample rows --------------------------------------------
    print("\n--Muestras de nombres separados (2006) --")
    df = pd.read_csv(
        os.path.join(CANDIDATOS_DEST, "candidatos_2006.csv"),
        encoding="utf-8-sig", dtype=str, nrows=10,
    )
    for _, row in df.head(5).iterrows():
        print(f"  {row['nombres']:<25s} | {row['apellido_paterno']:<15s} | {row.get('apellido_materno', ''):<15s}")

    print("\n--Muestras de alcaldes con votos (2018) --")
    df = pd.read_csv(
        os.path.join(ALCALDES_DIR, "alcaldes_con_votos_2018.csv"),
        encoding="utf-8-sig", dtype=str, nrows=10,
    )
    for _, row in df.head(5).iterrows():
        votos = row.get("total_votos", "")
        print(f"  {row['ubigeo']} {row['organizacion_politica'][:40]:<40s} votos={votos}")

    print("\nListo!")


if __name__ == "__main__":
    main()
