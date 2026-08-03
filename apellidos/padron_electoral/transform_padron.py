"""
Transform raw RENIEC padron electoral Excel files into clean, formatted Excel files.
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from copy import copy

# Configuration per year
CONFIGS = {
    2006: {
        'file': '2006_ERM.xlsx',
        'age_groups': [
            ('menores_20', 'Menores de 20'),
            ('20_39', '20 a 39'),
            ('40_69', '40 a 69'),
            ('70_mas', '70 a más'),
        ],
        # Data columns: 0-indexed positions in the row tuple
        # Total: 5,6,7  |  Age1: 8,9,10  |  Age2: 11,12,13  |  Age3: 14,15,16  |  Age4: 17,18,19
        'data_start': 5,
        'n_groups': 4,
    },
    2010: {
        'file': '2010_ERM.xlsx',
        'age_groups': [
            ('menores_30', 'Menores de 30'),
            ('30_59', '30 a 59'),
            ('60_mas', '60 a más'),
        ],
        'data_start': 5,
        'n_groups': 3,
    },
    2014: {
        'file': '2014_ERM.xlsx',
        'age_groups': [
            ('menores_30', 'Menores de 30'),
            ('30_59', '30 a 59'),
            ('60_mas', '60 a más'),
        ],
        'data_start': 5,
        'n_groups': 3,
    },
    2018: {
        'file': '2018_ERM.xlsx',
        'age_groups': [
            ('menores_30', 'Menores de 30'),
            ('30_59', '30 a 59'),
            ('60_mas', '60 a más'),
        ],
        'data_start': 5,
        'n_groups': 3,
    },
    2022: {
        'file': '2022_ERM.xlsx',
        'age_groups': [
            ('menores_30', 'Menores de 30'),
            ('30_59', '30 a 59'),
            ('60_mas', '60 a más'),
        ],
        'data_start': 5,
        'n_groups': 3,
    },
}


def parse_year(year):
    """Parse a raw RENIEC Excel file and return list of distrito-level dicts."""
    cfg = CONFIGS[year]
    wb = openpyxl.load_workbook(cfg['file'], read_only=True)
    ws = wb.active

    current_dept = None
    current_prov = None
    records = []

    for row in ws.iter_rows(min_row=7, values_only=True):
        # Identify row type based on which column has the name
        col_b = row[1]  # departamento
        col_c = row[2]  # provincia
        col_d = row[3]  # distrito
        col_e = row[4]  # ubigeo

        # Skip empty rows and footer rows
        if col_b is None and col_c is None and col_d is None:
            continue

        # Departamento row
        if col_b is not None and col_c is None and col_d is None:
            name = str(col_b).strip()
            if name.lower() in ('total', 'nacional', 'nacional '):
                continue
            current_dept = name
            continue

        # Provincia row
        if col_c is not None and col_d is None:
            current_prov = str(col_c).strip()
            continue

        # Distrito row
        if col_d is not None and col_e is not None:
            ubigeo_raw = str(col_e).strip()
            # Zero-pad to 6 digits
            ubigeo = ubigeo_raw.zfill(6)

            distrito = str(col_d).strip()
            ds = cfg['data_start']

            def safe_int(val):
                if val is None:
                    return 0
                try:
                    return int(round(float(val)))
                except (ValueError, TypeError):
                    return 0

            rec = {
                'ubigeo': ubigeo,
                'departamento': current_dept,
                'provincia': current_prov,
                'distrito': distrito,
                'total_electores': safe_int(row[ds]),
                'hombres': safe_int(row[ds + 1]),
                'mujeres': safe_int(row[ds + 2]),
            }

            # Age groups
            for i, (key, _label) in enumerate(cfg['age_groups']):
                offset = ds + 3 + i * 3
                rec[f'{key}_total'] = safe_int(row[offset])
                rec[f'{key}_hombres'] = safe_int(row[offset + 1])
                rec[f'{key}_mujeres'] = safe_int(row[offset + 2])

            records.append(rec)

    wb.close()
    return records


def build_columns(year):
    """Return ordered list of (key, header_label) for the output columns."""
    cfg = CONFIGS[year]
    cols = [
        ('ubigeo', 'Ubigeo'),
        ('departamento', 'Departamento'),
        ('provincia', 'Provincia'),
        ('distrito', 'Distrito'),
        ('total_electores', 'Total electores'),
        ('hombres', 'Hombres'),
        ('mujeres', 'Mujeres'),
    ]
    for key, label in cfg['age_groups']:
        cols.append((f'{key}_total', f'{label} - Total'))
        cols.append((f'{key}_hombres', f'{label} - Hombres'))
        cols.append((f'{key}_mujeres', f'{label} - Mujeres'))
    return cols


# -- Styling constants --
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

EVEN_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
ODD_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)

NUMBER_FORMAT = '#,##0'
TEXT_COLS = {'ubigeo', 'departamento', 'provincia', 'distrito'}


def format_worksheet(ws, columns, records, sheet_title=None):
    """Write records to a worksheet with professional formatting."""
    if sheet_title:
        ws.title = sheet_title

    # Write headers
    for col_idx, (key, label) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Write data
    for row_idx, rec in enumerate(records, 2):
        is_even = (row_idx % 2 == 0)
        fill = EVEN_FILL if is_even else ODD_FILL

        for col_idx, (key, _label) in enumerate(columns, 1):
            val = rec.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.fill = fill

            if key in TEXT_COLS:
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if key == 'ubigeo':
                    cell.number_format = '@'  # Force text
            else:
                cell.number_format = NUMBER_FORMAT
                cell.alignment = Alignment(horizontal='right', vertical='center')

    # Auto-filter
    last_col_letter = get_column_letter(len(columns))
    last_row = len(records) + 1
    ws.auto_filter.ref = f'A1:{last_col_letter}{last_row}'

    # Freeze panes (freeze header row)
    ws.freeze_panes = 'A2'

    # Adjust column widths
    for col_idx, (key, label) in enumerate(columns, 1):
        if key == 'ubigeo':
            width = 10
        elif key == 'departamento':
            width = 18
        elif key == 'provincia':
            width = 20
        elif key == 'distrito':
            width = 24
        elif 'total_electores' in key:
            width = 16
        else:
            # Measure header length, minimum 14
            width = max(14, len(label) + 2)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Set row height for header
    ws.row_dimensions[1].height = 30


def main():
    all_data = {}

    for year in [2006, 2010, 2014, 2018, 2022]:
        print(f'Parsing {year}...', end=' ')
        records = parse_year(year)
        columns = build_columns(year)
        all_data[year] = (records, columns)
        print(f'{len(records)} distritos')

        # Individual file
        out_file = f'padron_electoral_ERM{year}.xlsx'
        wb = openpyxl.Workbook()
        ws = wb.active
        format_worksheet(ws, columns, records, sheet_title=f'ERM {year}')
        wb.save(out_file)
        print(f'  -> {out_file}')

    # Consolidated file
    print()
    print('Creating consolidated file...')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for year in [2006, 2010, 2014, 2018, 2022]:
        records, columns = all_data[year]
        ws = wb.create_sheet(title=f'ERM {year}')
        format_worksheet(ws, columns, records)

    wb.save('padron_electoral_consolidado.xlsx')
    print('  -> padron_electoral_consolidado.xlsx')
    print()
    print('Done!')


if __name__ == '__main__':
    main()
