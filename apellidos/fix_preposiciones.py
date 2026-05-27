"""
fix_preposiciones.py
====================
Corrige los 584+ casos (alcaldes) y 6440+ (candidatos) donde apellido_paterno
quedó como preposición (DE, DEL, LA, LAS, LOS, EL) por la regla de separación
"últimas 2 palabras = apellidos".

Lógica (aplicada iterativamente para casos encadenados como DE LA CRUZ):
  1. apellido_materno = preposición + " " + apellido_materno actual
  2. apellido_paterno = última palabra del campo nombres
  3. nombres = todo lo anterior

Archivos afectados:
  - alcaldes_con_votos/alcaldes_con_votos_{2006,2010,2014}.csv
  - candidatos_completos/candidatos_{2006,2010,2014}.csv

Luego regenera los 5 Excel formateados.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PREPS = {"DE", "DEL", "LA", "LAS", "LOS", "EL"}


def fix_preposiciones(df):
    """Corrige filas donde apellido_paterno es una preposición.
    Aplica iterativamente para manejar casos encadenados (ej: DE LA CRUZ).
    Retorna (df_corregido, lista_de_cambios_para_reporte).
    """
    cambios = []
    df = df.copy()

    mask = df["apellido_paterno"].str.strip().str.upper().isin(PREPS)
    indices = df.index[mask].tolist()

    for idx in indices:
        orig_nombres = str(df.at[idx, "nombres"]).strip()
        orig_ap = str(df.at[idx, "apellido_paterno"]).strip()
        orig_am = str(df.at[idx, "apellido_materno"]).strip()

        nombres = orig_nombres
        ap = orig_ap
        am = orig_am

        # Aplicar iterativamente mientras apellido_paterno sea preposición
        while ap.upper() in PREPS:
            am = ap + " " + am  # prepend preposición al apellido materno
            partes = nombres.split()
            if len(partes) < 2:
                # No hay más palabras para extraer, dejar como está
                break
            ap = partes[-1]      # última palabra pasa a ser apellido paterno
            nombres = " ".join(partes[:-1])

        df.at[idx, "nombres"] = nombres
        df.at[idx, "apellido_paterno"] = ap
        df.at[idx, "apellido_materno"] = am

        cambios.append({
            "idx": idx,
            "antes_nombres": orig_nombres,
            "antes_ap": orig_ap,
            "antes_am": orig_am,
            "despues_nombres": nombres,
            "despues_ap": ap,
            "despues_am": am,
        })

    return df, cambios


# === Procesar CSVs ===
archivos = []
for year in [2006, 2010, 2014]:
    archivos.append((
        os.path.join(BASE_DIR, "alcaldes_con_votos", f"alcaldes_con_votos_{year}.csv"),
        f"alcaldes_con_votos_{year}"
    ))
    archivos.append((
        os.path.join(BASE_DIR, "candidatos_completos", f"candidatos_{year}.csv"),
        f"candidatos_{year}"
    ))

total_cambios = 0

for csv_path, label in archivos:
    print(f"\n{'='*60}")
    print(f"Procesando: {label}")
    print(f"{'='*60}")

    df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
    df.columns = [c.strip() for c in df.columns]

    df_fixed, cambios = fix_preposiciones(df)
    n = len(cambios)
    total_cambios += n
    print(f"  Casos corregidos: {n}")

    # Mostrar ejemplos (máximo 5)
    for c in cambios[:5]:
        print(f"  ANTES:   nombres={c['antes_nombres']:<35} ap={c['antes_ap']:<5} am={c['antes_am']}")
        print(f"  DESPUÉS: nombres={c['despues_nombres']:<35} ap={c['despues_ap']:<5} am={c['despues_am']}")
        print()

    # Verificar si quedaron preposiciones sin resolver
    residual = df_fixed["apellido_paterno"].str.strip().str.upper().isin(PREPS).sum()
    if residual > 0:
        print(f"  ⚠ Quedan {residual} casos sin resolver (nombres con 1 sola palabra)")
        problem = df_fixed[df_fixed["apellido_paterno"].str.strip().str.upper().isin(PREPS)]
        for _, row in problem.head(3).iterrows():
            print(f"    {row['nombres']} | {row['apellido_paterno']} | {row['apellido_materno']}")

    # Guardar
    df_fixed.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"  Guardado: {csv_path}")

print(f"\n{'='*60}")
print(f"TOTAL de correcciones: {total_cambios}")
print(f"{'='*60}")


# === Regenerar los 5 Excel ===
print("\n\nRegenerando los 5 archivos Excel...")

CSV_DIR = os.path.join(BASE_DIR, "alcaldes_con_votos")
YEARS = [2006, 2010, 2014, 2018, 2022]

HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
EVEN_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
CENTER_COLS = {"ubigeo", "sexo", "total_votos", "dni"}

for year in YEARS:
    csv_path = os.path.join(CSV_DIR, f"alcaldes_con_votos_{year}.csv")
    xlsx_path = os.path.join(BASE_DIR, f"alcaldes_{year}.xlsx")

    print(f"  Excel {year}...", end=" ")
    df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
    df.columns = [c.strip() for c in df.columns]

    df.to_excel(xlsx_path, index=False, sheet_name=f"Alcaldes {year}", engine="openpyxl")

    wb = load_workbook(xlsx_path)
    ws = wb.active
    num_cols = ws.max_column
    num_rows = ws.max_row
    headers = [ws.cell(row=1, column=c).value for c in range(1, num_cols + 1)]

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx in range(2, num_rows + 1):
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.fill = fill
            col_name = headers[col_idx - 1]
            if col_name and col_name.lower() in CENTER_COLS:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

    for col_idx in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(num_rows + 1, 1000)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        adjusted = min(max_len + 3, 45)
        ws.column_dimensions[col_letter].width = max(adjusted, 8)

    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows}"
    ws.freeze_panes = "A2"

    wb.save(xlsx_path)
    print(f"OK ({num_rows - 1} filas)")

print("\nListo!")
