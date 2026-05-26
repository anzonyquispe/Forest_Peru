"""
completar_sexo.py
=================
Completa el campo 'sexo' faltante en los CSVs de candidatos y alcaldes
para los años 2006, 2010 y 2014 usando clasificación basada en tokens
de nombres peruanos.

Luego regenera los 5 archivos Excel formateados.

Uso:
    python completar_sexo.py
"""

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ALCALDES_DIR = os.path.join(BASE_DIR, "alcaldes_con_votos")
CANDIDATOS_DIR = os.path.join(BASE_DIR, "candidatos_completos")

# =====================================================================
# GENDER TOKEN DICTIONARIES
# =====================================================================
# Comprehensive sets of name tokens mapped to gender.
# Built from analysis of Peruvian naming conventions including
# Spanish, Quechua, and phonetic variants common in Peru.

FEMALE_TOKENS = {
    # --- A ---
    "ABIGAI", "ADALI", "ADASELI", "ADELHI", "ADELI", "ADIE", "ADILDE",
    "AGDI", "AGIE", "AHIDEE", "AHIDELI", "AIDEE", "AIDELI", "AIMME",
    "ALCIONE", "ALDEYDE", "ALENI", "ALHELI", "ALILI", "ALIOSI",
    "ALISMENI", "ALSIONE", "AMAINE", "AMELI", "ANACE", "ANACELI",
    "ANADEYSI", "ANAI", "ANALI", "ANANI", "ANDRIETTE", "ANELI", "ANELU",
    "ANELVI", "ANGEE", "ANGELI", "ANGELUZZI", "ANGGIE", "ANILU", "ANITE",
    "ANNELIESSE", "ANNELINE", "ANNYE", "ANTUANE", "ANYHILI", "ANYI",
    "ARACELLI", "ARANZASU", "ARASELI", "ARAZELI", "ARBELI", "ARCELI",
    "ARELI", "ARLENI", "ARLETTI", "ARNELI", "ARYHERI", "AYDDE", "AYDE",
    "AYDEE", "AYDELI", "AYDYE", "AYEYE", "AYIDE", "AYLI", "AYME",
    "AYMEE", "AYRUNEDI", "AZIZI",
    # --- B ---
    "BABERLI", "BABI", "BANESI", "BANI", "BARINI", "BASTI", "BEANI",
    "BECI", "BECSI", "BEERI", "BEHTI", "BELCI", "BELQUI", "BELSI",
    "BELU", "BERENISSE", "BERI", "BERLLI", "BERSABE", "BETCI", "BETHI",
    "BETHZABE", "BETSABE", "BETSABEE", "BETSAI", "BETSAIDE", "BETTSABE",
    "BETZABE", "BETZI", "BEVELU", "BEXABE", "BEXI", "BEYVI", "BEYZAVE",
    "BIANE", "BIQUE", "BLENI", "BLICE", "BONNEE", "BRENDESI",
    "BRIGGITE", "BRIGGITHE", "BRIGGITTE", "BRIGUITHE", "BRINDISI",
    # --- C ---
    "CAMURI", "CANI", "CAROLAI", "CATERINE", "CATHERYNE", "CATIALAINI",
    "CATTI", "CECI", "CELI", "CENI", "CESI", "CHACHI", "CHALI", "CHANI",
    "CHELI", "CHESI", "CHESSIE", "CICELI", "CISE", "CLARI", "CLAYRE",
    "CLAYRI", "CLEDI", "CLEIDI", "CLENI", "CLEOFFE", "CLEOTILDE",
    "CLERI", "CLESI", "CLEYDI", "CLIDE", "CLISFE", "CORALI", "COSSI",
    "COTI", "CRAICE", "CRENE", "CRONI", "CRYSTI", "CUSI",
    # --- D ---
    "DAGLI", "DAGNE", "DAGNI", "DAIMI", "DAISE", "DAISI", "DAIZI",
    "DALMACE", "DAMARI", "DANAE", "DANEE", "DANESI", "DANITZE",
    "DARLENI", "DARLI", "DARMI", "DARNI", "DARVE", "DAYCE", "DAYCI",
    "DAYESE", "DAYLI", "DAYNE", "DAYSE", "DAYSI", "DAYSSI", "DECI",
    "DEILE", "DEINE", "DEISE", "DEISI", "DEISSI", "DELCE", "DELCI",
    "DELEINE", "DELFI", "DELHI", "DELI", "DELMI", "DELSI", "DELVI",
    "DENCE", "DENEIRE", "DENESSI", "DENISSE", "DENNESI", "DENNISE",
    "DENNISSE", "DEONI", "DERI", "DERMALI", "DERSI", "DEXCI", "DEXI",
    "DEYBI", "DEYCI", "DEYLI", "DEYMI", "DEYNI", "DEYSE", "DEYSERLI",
    "DEYSI", "DEYSSI", "DHANI", "DHIDMI", "DIAMILE", "DIANETHI",
    "DIANIRE", "DIARELI", "DICKI", "DILCE", "DILME", "DIOMI", "DIONE",
    "DIONI", "DIOTILDE", "DIXI", "DOILI", "DOLI", "DONATILDE", "DORALE",
    "DORALI", "DORSI", "DOYLI", "DRIAME", "DRISSI", "DUANI", "DUNE",
    "DUVI", "DYANE", "DYNISSE", "DYOVI",
    # --- E ---
    "EDDE", "EDILENI", "EDIYULVE", "EDME", "EDMEE", "EDMI", "EDOLI",
    "EIDI", "EIMI", "ELCI", "ELEILDE", "ELEINE", "ELEVI",
    "ELINCE", "EMABILE", "EMELI", "EMELINNE", "EMILCE", "EMILENE",
    "EMILSE", "EMYLSE", "ENGGIE", "ENI", "ENILDE", "ENOE", "EONISE",
    "ERENE", "ERIOTILDE", "ERLI", "ESMENI", "ESTEFANI", "ESTEFANIE",
    "ESTEFFANI", "ESTEFFANIE", "ESTEPHANI", "ESTERLI", "ESTHEFANI",
    "ESTANI", "EUDE", "EUNESI", "EUNICIE", "EUNISE", "EVERLU",
    "EVOLINE", "EYDE", "EYDI", "EYDIE", "EYNI",
    # --- F ---
    "FALCONERI", "FARIDE", "FEBE", "FEBI", "FENI", "FEVI", "FLAVI",
    "FLORISVELI", "FRANCELI", "FRANSINE", "FREYSI", "FRIDE",
    # --- G ---
    "GADALI", "GAVI", "GEAMNI", "GEANETTE", "GEIBI", "GEIRI", "GELI",
    "GEMALLI", "GENI", "GEORDAMI", "GERALDINE", "GERALDYNE",
    "GERALINNE", "GERDE", "GERSI", "GESSI", "GEYDI", "GEYMI", "GEYSI",
    "GIANNINE", "GICELI", "GIEZI", "GINNI", "GISELE", "GLADI", "GLENI",
    "GLERI", "GLEYCI", "GLEYDI", "GLIDE", "GLOTILDE", "GLUDE", "GOLDI",
    "GRATELI", "GREASE", "GREISI", "GREYSHI", "GREYSI", "GREYSSI",
    "GUEIDI", "GUISSELLE", "GUISELLE",
    # --- H ---
    "HAIDEE", "HAYDE", "HAYLE", "HELDI", "HELENISE", "HELINE", "HERI",
    "HERMANI", "HERSI", "HEVELI", "HEVERLI", "HEYDE", "HEYDI", "HOYSI",
    "HUBI",
    # --- I ---
    "IBI", "IBONI", "IDE", "IDEYVI", "ILSI", "INGLI", "INGRI", "INOE",
    "IOHANNIE", "IRENI", "IRETTE", "IRONI", "ISMELI", "ISOE",
    "ISRAELI", "IVE", "IVETE", "IVETHSI", "IVONNI",
    # --- J ---
    "JACKELINNE", "JACKELYNE", "JACKELYNNE", "JACKIE", "JAHIDE",
    "JAIDE", "JAINE", "JAKELINE", "JAKELINEE", "JAKELYNE", "JAKILINI",
    "JANCILEIDE", "JANELSI", "JAQUELINNE", "JAQUELYNE", "JAQUELYNNE",
    "JASCELLI", "JAYDI", "JAYU", "JEANNI", "JEHILEE", "JEINE", "JEINI",
    "JELSI", "JENAI", "JENIREE", "JENMI", "JERLI", "JESI", "JESIE",
    "JEYDI", "JEYMI", "JEYNI", "JHANISE", "JHAQUELINE", "JHAYDEE",
    "JHENNI", "JHENSI", "JHEYMI", "JHEYSI", "JHOSELINE", "JHOSELYNE",
    "JHOYSI", "JHUDI", "JOAMI", "JOHSELINNE", "JOOISI", "JOSELINNE",
    "JOSSIANE", "JOSSIE", "JOYCI", "JOYCIE", "JOYSI", "JOYSIE",
    "JUANI", "JULEYSI", "JYNNI",
    # --- K ---
    "KACELI", "KACTHERINE", "KADDYE", "KARELI", "KARLEINE", "KAROLE",
    "KATARINE", "KATERI", "KATERINE", "KATERYNE", "KATHERINNE",
    "KATMERINE", "KATTERINE", "KATTHERINE", "KATTI", "KEELI", "KEILI",
    "KELI", "KEMI", "KENNE", "KETTI", "KEYDI", "KEYLE", "KEYLI",
    "KEYSI", "KEYSSI", "KILDARE", "KLAYRE", "KUNTI",
    # --- L ---
    "LACENI", "LADI", "LADYDI", "LAINETTE", "LEAMERSI", "LEBI", "LEIDE",
    "LEIDI", "LEINE", "LEISI", "LEISLI", "LEISLIE", "LENCI", "LENDI",
    "LENNI", "LESBI", "LESGIE", "LESLYE", "LESSLIE", "LESVI", "LESWI",
    "LEXMI", "LEYBI", "LEYDI", "LEYNI", "LEYRI", "LEYSI", "LEYVI",
    "LIANI", "LIBNI", "LICI", "LIDANI", "LIDI", "LILE", "LINNE",
    "LIRPE", "LISBETTE", "LISSETE", "LISVE", "LISVETTE", "LITZI",
    "LIVE", "LIVIDI", "LIZI", "LIZZETE", "LIZZETTE", "LLAINI", "LLANELI",
    "LLANI", "LLANINI", "LLARLI", "LLELI", "LLEME", "LLENI", "LLENSI",
    "LLERI", "LLERME", "LLEYMI", "LLEYSI", "LLIANCI", "LLIMI",
    "LLIMILINE", "LLINE", "LLOHANI", "LLONI", "LLORVI", "LLOVANI",
    "LLOYSI", "LLULI", "LLUNELI", "LOANI", "LOYDE", "LUANI", "LUDANI",
    "LUDI", "LURDE", "LURI", "LUSI", "LUZDARI", "LUZLI", "LUZMERI",
    "LYLI",
    # --- M ---
    "MADAI", "MADALEYNE", "MADALEYNI", "MADAME", "MADDAI", "MADELAYNE",
    "MADELEINI", "MADELEYDE", "MADELEYDI", "MADELEYNE", "MADELEYNI",
    "MADELI", "MADILENI", "MAGALLI", "MAGDALI", "MAGDE", "MAGE",
    "MAGGALI", "MAGGI", "MAGHI", "MAHLI", "MAHUMI", "MALENI", "MALHI",
    "MALLURI", "MANAI", "MANIE", "MAPI", "MARBEYLI", "MARBI", "MARDELI",
    "MARELI", "MARFI", "MARIALI", "MARIANELI", "MARIBI", "MARICE",
    "MARILE", "MARILEIDE", "MARILI", "MARILOLI", "MARINELI", "MARIOLI",
    "MARISCELI", "MARISELI", "MARISU", "MARITERE", "MARIUXI",
    "MARIYIELE", "MARIZE", "MARIZELI", "MARLE", "MARLENI", "MARLENNI",
    "MARLESI", "MARLHENI", "MARLUBE", "MARLUVE", "MARSHELLI", "MARUVENI",
    "MARYLU", "MARYOLI", "MARYORI", "MARYORIE", "MARYURI", "MATISEE",
    "MAVI", "MAYANI", "MAYBEE", "MAYDELENI", "MAYELI", "MAYLENE",
    "MAYLI", "MAYLIE", "MAYNE", "MAYORI", "MAYRENI", "MAYUMHI",
    "MAYYOLI", "MECHE", "MECHI", "MEDALI", "MEDELI", "MEDI", "MEE",
    "MEKE", "MELANNIE", "MELDINE", "MELEDI", "MELENE", "MELI",
    "MELISSE", "MELODI", "MELSI", "MELVI", "MENNI", "MERCI", "MERE",
    "MERICE", "MERUPE", "MEYDI", "MEYSI", "MIAMI", "MIANI", "MICHEELE",
    "MIGUESI", "MIKSI", "MILDRE", "MILEIDE", "MILENE", "MILENI",
    "MILEYDE", "MILTRE", "MILU", "MIREILI", "MIRELE", "MIRELI",
    "MIRLENE", "MIRLI", "MIROPE", "MIRUSDI", "MITILDE", "MITTI",
    "MITTZE", "MIXI", "MIXU", "MOSOLINE", "MUNIRE", "MYLAYDE", "MYLENE",
    "MYLENI",
    # --- N ---
    "NAHOMI", "NAI", "NAIROBI", "NALDI", "NANCE", "NANSI", "NARDI",
    "NARI", "NARVI", "NATHALI", "NECKI", "NEFRETERI", "NEIDI", "NEISI",
    "NEISME", "NELCI", "NELENNI", "NELLYLENI", "NELSI", "NELVI", "NEPSI",
    "NERELI", "NERLI", "NESEI", "NESSI", "NETSI", "NEUCI", "NEVI",
    "NEXI", "NEYCI", "NEYDE", "NEYDI", "NEYLI", "NEYMI", "NEYRI",
    "NEYSENI", "NEYSI", "NEYSME", "NEYVI", "NHANCI", "NIBELI", "NICELLI",
    "NICOLI", "NIERE", "NILENE", "NITSI", "NOELI", "NOIMI", "NOMI",
    "NORAIMI", "NORELLI", "NOYMI", "NURE", "NYLSE",
    # --- O ---
    "OKEMI", "OLGUI", "ORIELE", "OVELI",
    # --- P ---
    "PATME", "PEGGI", "PEGUI", "PRELI", "PREYCI",
    # --- Q ---
    "QUELI", "QUENDI", "QUENI", "QUIMI",
    # --- R ---
    "RAEME", "RAYDEE", "RAYDI", "RAYME", "REYDE", "REYNE", "REYNERI",
    "REYNI", "RINCCI", "RISBELI", "RODE", "RODI", "RODMI",
    "ROICE", "ROMELI", "ROSALI", "ROSALU", "ROSANI", "ROSBELI",
    "ROSEE", "ROSELBI", "ROSELE", "ROSELI", "ROSEYVI", "ROSILU",
    "ROSMARI", "ROSMELI", "ROSMERE", "ROSMERI", "ROSMI", "ROSSE",
    "ROSSI", "ROSSLI", "ROSSMERI", "ROSTAI", "ROSTMERI", "ROSYRENE",
    "ROYSI", "RUBELINE", "RUSMERI", "RUTHALI", "RUTMELI", "RUVI",
    "RUVYE", "RUYDE", "RUZMERI",
    # --- S ---
    "SABI", "SAHIBI", "SAHORI", "SALUSTRE", "SALVI", "SANI",
    "SARAHI", "SARMELI", "SELENI", "SELENNE", "SELI", "SETNE", "SHANDE",
    "SHARMELI", "SHENE", "SHEYLI", "SHIRLE", "SHIRLLE", "SICI", "SILENI",
    "SILI", "SINARSHYI", "SINDI", "SINGRI", "SIRLENE", "SMAREDRE",
    "SOLANGI", "SOLI", "SORI", "SSISI", "STEFI", "STEISI", "STEPHANI",
    "STEYSI", "STHEFANI", "STTYFANNI", "SUCI", "SUGEI", "SUGEILI",
    "SUGELI", "SUJEI", "SUJEYDI", "SULE", "SULI", "SULMI",
    "SUNDI", "SUSELE", "SUSETI", "SUWOMI", "SYBILE",
    # --- T ---
    "TATI", "TECHI", "TESI", "THAYRI", "TRILCE",
    # --- V ---
    "VAITIARE", "VASTI", "VELSI", "VELU", "VERENICE", "VERENISSE",
    "VETZABEE", "VEXCI", "VIQUE", "VIQUI", "VIUHNNI",
    # --- W ---
    "WONIE",
    # --- X ---
    "XAI", "XIOMI", "XOMARIE",
    # --- Y ---
    "YABELI", "YACKELINE", "YACKILLINE", "YACQUELINE", "YADI", "YAIDE",
    "YAIME", "YAIMI", "YAKELI", "YAKELINE", "YAKELLINE", "YAKELLYNE",
    "YALI", "YAMALI", "YAMELI", "YAMERI", "YAMILI", "YAMILLE", "YANALI",
    "YANDI", "YANE", "YANELE", "YANELI", "YANETI", "YANETTE", "YANILDE",
    "YANINI", "YANIRE", "YANNE", "YANNI", "YAQUE", "YAQUELI",
    "YAQUELINE", "YAQUI", "YARESSI", "YARI", "YARIFE", "YARSELI",
    "YASARI", "YASERI", "YASMI", "YASUMI", "YAVE", "YAVELI", "YBI",
    "YEIMI", "YEINE", "YEISE", "YEISI", "YEIVI", "YELE", "YELI",
    "YELSI", "YELVI", "YEMBI", "YEMI", "YEMILE", "YENCI", "YENE",
    "YENI", "YENNE", "YENNI", "YERALDINE", "YERCI", "YERENI", "YERI",
    "YERTI", "YESI", "YESSELI", "YESSI", "YESSILE", "YETI", "YEYMI",
    "YEYSI", "YHASMINE", "YHASMIRE", "YHONNE", "YHOVI", "YIENE", "YIME",
    "YINELZIE", "YIRLE", "YOANE", "YOANI", "YOBALI", "YOBANE", "YOCELI",
    "YOCELINE", "YODI", "YOHANE", "YOJANI", "YOJENI", "YOLANE", "YOLE",
    "YOLENI", "YOLI", "YOLVI", "YORSI", "YOSDI", "YOSELI", "YOSELINE",
    "YOSELVE", "YOSHME", "YOSI", "YOSILU", "YOSSELINE", "YOVANI",
    "YOVANNI", "YOWALI", "YOYCI", "YOYSI", "YRENE", "YSTIVI", "YSVETTE",
    "YUBANE", "YUBERLI", "YUCELI", "YUDDI", "YUDE", "YUDI", "YUJAME",
    "YULE", "YULEISI", "YULEYCI", "YULEYSI", "YULINE", "YUNE", "YUNELI",
    "YURAICI", "YURIKU", "YUSELI", "YUSLI", "YUSSI", "YUVELE", "YUVI",
    "YVONE",
    # --- Z ---
    "ZADI", "ZARAI", "ZAYURI", "ZELENI", "ZELI", "ZELMI",
    "ZELYAMALI", "ZIMI", "ZOCI", "ZOIDELI", "ZOMELI", "ZONALI", "ZUELI",
    "ZUKI", "ZULI", "ZULLI", "ZULVE", "ZUNILDE", "ZURI", "ZURISADAI",
    "ZUSI", "ZUSSI",
    # Common female second-name tokens
    "ELIZABETH", "MARIBEL", "MARIA", "ROSA", "ROSARIO", "PILAR",
    "CARMEN", "JULIA", "JUANA", "ESTHER", "MARTHA", "ISABEL",
    "LOURDES", "YOLANDA", "GLORIA", "ROCIO", "MARITZA", "ROXANA",
    "MONICA", "PATRICIA", "BEATRIZ", "ALICIA", "VIRGINIA", "JOSEFINA",
    "VERONICA", "ESPERANZA", "DORIS", "MIRIAM", "VICTORIA", "TERESA",
    "LUISA", "ELENA", "IRENE", "LEONILA", "LEONOR", "GRACIELA",
    "HERMINIA", "NATIVIDAD", "SOLEDAD", "SATURNINA", "TOMASA",
    "PRUDENCIA", "OTILIA", "FILOMENA", "AVELINA", "BERTHA", "MARILU",
    "MARISOL", "NOEMI", "VANESSA", "VANESA", "KARINA", "MAGALY",
    "MILAGROS", "YESENIA", "FIORELLA", "JACKELINE", "JAQUELINE",
    "LILIANA", "ROSALIA", "CECILIA", "SUSANA", "DOLORES", "IRMA",
    "ELVIRA", "EMILIA", "EDITH", "ESMERALDA", "HILDA", "OLINDA",
    "NILDA", "CLEMENCIA", "CLEMENTINA", "LUCILA", "PAULINA", "ANTONIA",
    "ADRIANA", "EMPERATRIZ", "LEONILDA", "CLORINDA", "ESCOLASTICA",
    "GERTRUDIS", "HILARIA", "CRISTINA", "YOVANA", "JOVANA", "MARLENE",
    "CAROLINA", "TATIANA", "PAMELA", "DIANA", "KATHERINE", "JESSICA",
    "LISBETH", "LIZBETH", "ERIKA", "GIOVANNA", "KATHIA", "YANINA",
    "YANET", "ROMELIA", "SABINA", "ANGELICA", "MARCELA", "LUZMILA",
    "MARGARITA", "GREGORIA", "MERCEDES",
}

MALE_TOKENS = {
    # --- A ---
    "ABERLI", "ADME", "ADMIRI", "ADONI", "AKUMPARI", "ALBERTI",
    "ALBERTINI", "ALEXSI", "ALFIERI", "ALFREDI", "ALLENDE", "AMARANTE",
    "AMARU", "ANRRI", "ANTHONI",
    "ARCE", "ARPI", "ARRI", "AUREI",
    # --- B ---
    "BARI", "BARTOLOME", "BECQUE", "BELAUNDE", "BENE", "BENETTI",
    "BERLI", "BERNABE", "BERNAVE", "BERTONE", "BERTONI", "BIKE",
    "BILE", "BOCQUINI", "BORGE", "BRANLI", "BRAUNE", "BRUCELEE",
    "BRUCELI", "BRUNEI",
    # --- C ---
    "CAHUIDE", "CARLE", "CARLESI", "CARLESSI", "CARLEYLE", "CASELI",
    "CEDE", "CESARI", "CHAQUE", "CHARLE", "CHE", "CHEDLI", "CHERSI",
    "CLEME", "CLEUFE", "CLIFE", "CLINE", "CONSTANTE", "CONVERTTI",
    "CORCINU", "CREMLI",
    # --- D ---
    "DANDI", "DANINI", "DAQU", "DAVI", "DE", "DEIBI", "DENNI", "DERLI",
    "DEYBI", "DEYVI", "DIAGNE", "DIANME", "DIETTCHE", "DOANE", "DONI",
    "DUALTIME", "DUBERLI",
    # --- E ---
    "EDDYE", "EDERLI", "EDOU", "EDU", "EDUARTE", "EDUBARLI", "EILANDE",
    "EKI", "ELEUDU", "ELINCE", "ELIU", "ELIZALDE", "ELKI", "ELQUI",
    "ELWI", "EMERARIE", "EMILSE", "ENARTE", "ENDE", "ENRRIQUE",
    "ERASMI", "ERELI", "ERME", "ERNAU", "ESAU", "ESPIRITU",
    "ESTANISLAU", "ESTENI", "ETNI", "EUDI", "EURPEPE", "EYGAINE",
    # --- F ---
    "FALCONERE", "FALCONERI", "FALCONI", "FALCONIERE", "FALCONIERI",
    "FALIONERI", "FERNANDINI", "FERNI", "FIESHE", "FOCILI", "FORE",
    "FRANKYE", "FRASNELLI", "FREDELI", "FREI", "FREYDE", "FREYRE",
    # --- G ---
    "GALVANI", "GANDI", "GASMANI", "GEHU", "GELVI", "GENRI", "GENRRI",
    "GENRRIE", "GEOVANI", "GERVI", "GILDE", "GIMI", "GIOVANI",
    "GIRALDE", "GOETHE", "GORGE", "GRAZZIANI", "GUEMBI", "GUENI",
    "GUEYVI", "GUIME", "GUIOVANI",
    # --- H ---
    "HAMMERLI", "HAVE", "HENRRI", "HERBE", "HERLI", "HERLINE",
    "HERNANI", "HERNOLDE", "HIDE", "HILBE", "HILDERME", "HILQUE",
    "HIRUYUKI", "HODE", "HODI", "HOOBE", "HORAYME", "HOWI", "HUBERLI",
    "HUGUE", "HUILE", "HULORTIGUE", "HUMBE",
    # --- I ---
    "IDONE", "IHUAQUI", "IKE", "INDORFE", "INOCENTE", "IRIARTE",
    "ISAU", "IVANOE",
    # --- J ---
    "JABI", "JAEME", "JANRRI", "JARLI", "JAURE", "JAVI",
    "JEAMPIERRE", "JEFTE", "JEHU", "JENDRI", "JENRE", "JENRI",
    "JENRRI", "JEORGE", "JHIME", "JHOLE", "JHONI", "JHONNI",
    "JHOOVANI", "JHOSE", "JHOVANI", "JHOWVANI", "JIANI", "JIME",
    "JOLVI", "JORSI", "JOSEI", "JOSELI", "JOSFE", "JOSHADI", "JOSOE",
    "JOSUECI", "JOSUHE", "JOVANI", "JOVANNE", "JULE",
    # --- K ---
    "KAMIJU", "KENI", "KENNI", "KENYI", "KIKE", "KIQUE", "KOQUE",
    "KOSME", "KUBULKI", "KUISHINQUE", "KUKI",
    # --- L ---
    "LAI", "LEOHNE", "LEONARDI", "LEVINE", "LINDORFE", "LISNE",
    "LOARTE", "LOERE", "LOLI", "LORENSE", "LUATANI", "LUBECKE",
    "LUCIANI", "LUGIE", "LUIGGI",
    # --- M ---
    "MAGDI", "MAGHDI", "MALVI", "MANFRI", "MANRIQUE", "MANYARE",
    "MANZONI", "MARFORI", "MARZOLINE", "MELBI", "MELENDRE", "MERARI",
    "MERE", "MONGE", "MORANTE",
    # --- N ---
    "NAYME", "NEDDI", "NEFE", "NENE", "NEPHATALI", "NEPHTALI", "NEPTALI",
    "NIERE", "NIKYOLI", "NILVE", "NINIVE", "NIQUE", "NIRE", "NOBOU",
    "NOLI", "NORBI", "NORJE",
    # --- O ---
    "OBE", "OFNI", "OFRANI", "OJANI", "OLIVIE", "ONOFRE", "ORLANDINE",
    "ORLANDINI",
    # --- P ---
    "PABLIKOWESKI", "PACSI", "PAJONE", "PAPI", "PARMI", "PELE",
    "PERCI", "PERSI", "PERSSI", "PERUSE", "PIERE", "POLI", "PONCE",
    "PRIALE", "PRILE",
    # --- Q ---
    "QUELVE", "QUENIDE",
    # --- R ---
    "RAI", "RAMONETE", "RANDHU", "RANFORTE", "RASURE", "REBOLI",
    "RELE", "RELI", "RENIDE", "REYDI", "REYNEE", "REYNERI", "RICARTE",
    "RICKE", "RIMBERTE", "RIMBERTI", "RIQUE", "RIQUELME", "ROBESPIERRE",
    "ROCE", "ROCHE", "ROCINI", "RODI", "ROE", "ROGGI", "ROGNE",
    "ROLE", "ROLFE", "ROMANI", "ROMNI", "RONALDE", "RONDAU",
    "RONE", "RONIE", "RORI", "ROSEBELTE", "ROSELI", "ROWIE",
    "ROYNE", "RUDE", "RUFME", "RULE", "RULI", "RULLI", "RURU",
    # --- S ---
    "SABU", "SAJDI", "SANDRI", "SANTONI", "SEMELI", "SHANTU",
    "SI", "SILBESTRE", "SINAI", "SIVESTRE", "SOANE", "SUANE", "SUDANI",
    # --- T ---
    "TANU", "TEDI", "TEOLU", "TURIKI", "TURIPI",
    # --- U ---
    "UAQUI", "UGARTE", "ULBE", "ULDE", "URIBE",
    # --- V ---
    "VENCI", "VENNE", "VERCELLI", "VERSELI", "VIJAI", "VIKE",
    "VISENTE", "VIVANCLI",
    # --- W ---
    "WALQUI", "WANPEAU", "WENSESLAU", "WENSISLAU", "WESLI", "WILDE",
    "WILE", "WILI", "WILSE", "WISENI", "WITHLE", "WU", "WUILE",
    "WUILLI",
    # --- Y ---
    "YASALDE", "YASMANI", "YAVE", "YENCI", "YENERE", "YENSI", "YERTE",
    "YHONI", "YLDE", "YIMI", "YIMME", "YNOCENTE", "YOBANI",
    "YOCHI", "YOE", "YOFRE", "YOHNI", "YONCLE",
    "YONE", "YONHE", "YONNI", "YORDI", "YORI", "YOSE", "YOVANI",
    "YSAI", "YSOE", "YUDZI",
    # --- Z ---
    "ZHU",
    # Common male second-name tokens
    "CARLOS", "PEDRO", "JOSE", "JUAN", "LUIS", "JORGE", "FRANCISCO",
    "ALBERTO", "ANTONIO", "ENRIQUE", "FELIPE", "MANUEL", "MIGUEL",
    "RAFAEL", "RICARDO", "ROBERTO", "HECTOR", "CESAR", "FELIX",
    "MARIO", "VICTOR", "ARTURO", "EMILIO", "GREGORIO", "LEOPOLDO",
    "MOISES", "NELSON", "OSCAR", "PERCY", "RAMON", "RAUL", "ROGER",
    "ROMAN", "RUBEN", "SALOMON", "SANTOS", "SEGUNDO", "SIMON",
    "TEODORO", "TEODOSIO", "TEOFILO", "ULDARICO", "WALTER", "WILLIAM",
    "ZENOBIO", "VIRGILIO", "BENIGNO", "FORTUNATO", "ELIAS", "DARIO",
    "LEONIDAS", "PLACIDO", "PELAYO", "ADOLFO", "CELESTINO", "JUSTINIANO",
    "ANIBAL", "NICASIO", "GILBERTO", "MODESTO", "ABEL", "EDGAR",
    "HENRY", "MARTIN", "DANIEL", "MARCOS", "DAVID", "MICHAEL",
    "RONALD", "ALEXANDER", "DENNIS", "JHONATAN", "ROLANDO",
    "ROBINSON", "CRISTOBAL", "JACOB", "JOEL", "MAURICIO", "ALONSO",
    "GUSTAVO", "HERNAN", "ARMANDO", "PASCUAL", "ALFONSO", "VALENTINO",
    "VALENTIN", "GAUDENCIO", "AGAPITO", "VICENTE", "WILFREDO",
    "LEONCIO", "GODOFREDO", "FILOMENO", "LAZARO", "ANANIAS",
    "JENRRY", "JHON", "WILMER", "JONATHAN", "ANDERSEN",
    "FABRICIO", "PAOLO", "ANTHONY", "ADRIAN", "CHRISTIAN",
    "NIKOLAI", "ESTALIN", "DARBIN",
}

# Specific overrides for ambiguous names where context or the full
# compound name determines gender differently from what tokens suggest
OVERRIDE_MAP = {
    # Names where first token appears ambiguous but full name is clear
    "ALACOQUE ALBERTO": "M",
    "AYME GARY": "M",
    "DEYBI TATIANA": "F",
    "DEYBI ALFONSO": "M",
    "DEYBI ANTONIO": "M",
    "DEYBI LUIS": "M",
    "DE JOEL HERNAN": "M",
    "DE LA CRUZ": "M",
    "DE LA PAZ": "M",
    "DIONE OMAR": "M",
    "DIONE MAYRA": "F",
    "DIONE NOEMI": "F",
    "DIONE RUTH": "F",
    "DIONI RADI": "M",
    "DIONI MARIA": "F",
    "E ELISA": "F",
    "E FELIX": "M",
    "ELINCE AMELLINO BAILON DE": "F",
    "ELINCE DELFIN": "M",
    "ELKI HAYDEE": "F",
    "ESPIRITU CENCIA": "F",
    "ESPIRITU DE LA": "F",
    "ESPIRITU JUDY": "F",
    "GEOVANI TANNIA": "F",
    "GEOVANI ESTHER": "F",
    "GIOVANI ESTHER": "F",
    "GIOVANI EDILBERTO": "M",
    "GIOVANI MAURICIO": "M",
    "HAYLE CATHERYNE": "F",
    "HAYLE SALECIO": "M",
    "HEYDE ODO": "M",
    "HEYDE GISELA": "F",
    "INOCENTE TEODORO POLO DE": "F",
    "JOVANI": "M",
    "LLONI MARLO": "M",
    "LLONI SUELPEREZ RAMIREZ": "F",
    "LLONNI GABINO": "M",
    "LOLI EPIFANIO": "M",
    "LOLI FERNANDO": "M",
    "LOLI FROILAN": "M",
    "LOLI MANSUETO": "M",
    "LOLI MAXIMO": "M",
    "LOLI RECCI": "F",
    "LOLI RUBEN": "M",
    "NENE CLYDE": "M",
    "NENE LUIS": "M",
    "NIERE CASILDA": "F",
    "RODI MARTHA": "F",
    "RODI AUDENCIO": "M",
    "RODI CELSO": "M",
    "RODI RIVER": "M",
    "ROSELI AMERICO": "M",
    "ROSELI MARILYN": "F",
    "ROSSI BETTY": "F",
    "ROSSI DAVID": "M",
    "ROSSI EDUARDO": "M",
    "ROSSI ELIZABETH": "F",
    "ROSSI GALVAN": "M",
    "ROSSI INALDO": "M",
    "ROSSI MARILU": "F",
    "ROSSI MARY": "F",
    "ROSSI MELVA": "F",
    "ROSSI NELIDA": "F",
    "ROSSI PILAR": "F",
    "ROSSI RODRICH": "M",
    "ROSSI RUDITA": "F",
    "ROSSI SALOME": "F",
    "ROSSI VILMA": "F",
    "ROSSI YENETH": "F",
    "ROSSI YOVANNA": "F",
    "RULI JAIDI": "F",
    "RULI EVERT": "M",
    "RULI ROBERTO": "M",
    "RODE ELIZABETH": "F",
    "RODE KELITA": "F",
    "YAVE HERLINDA": "F",
    "YAVE ISAIAS": "M",
    "YEMI GREGORIO": "M",
    "YENCI WILFREDO": "M",
    "YENCI LEISLI": "F",
    "YENSI EUSEBIO": "M",
    "YOBANI DAVID": "M",
    "YOBANI ISABEL": "F",
    "YOVANI ADOLFO": "M",
    "YOVANI DEL PILAR": "F",
    "YOVANI EMMI": "F",
    "YOVANI ESMITH": "F",
    "YOVANI GUSVINDA": "F",
    "YOVANI MAYLENE": "F",
    "YOVANI MELISA": "F",
    "YOVANI YANINA": "F",
    "YOVANI YAQUELI": "F",
    "YOLVI ESPINOZA LOPEZ": "M",
    "YOLVI GUSTAVO": "M",
    "YOLVI HERNAN": "M",
    "YOLVI MAGALY": "F",
    "YOLVI MARILU": "F",
    "RONE YUBIT SANTA": "F",
    "RONE WILMER": "M",
    "CLIDE SANTOS": "M",
    "REYNERI LIDUVINA": "F",
    "REYNERI SANTIAGO": "M",
    "DONAIRE DOMINGUEZ": "M",
    "BERNABE HILDA": "F",
    "BERNABE LUPITA": "F",
    "CELI GALINDO": "M",
    "CELI SAUL": "M",
    "CELI MAGALI": "F",
    "CELI ROSAURA": "F",
    "CELI YAMALI": "F",
    "GENI LLERME": "F",
    "GENI YONNY": "M",
    "DENNI NERI": "F",
    "DENNI NESTOR": "M",
    "DENI DARSHAM": "M",
    "MERCI YOVANI": "M",
    "PAPI PATROCINIO": "M",
    "ARELI UZIEL": "M",
    "ANGUI JESUS": "M",
    "JEYVI RUDVI": "M",
    "LLI JHONATAN": "M",
    "LEIHGABE": "M",
    "MIZRAI": "M",
    "RAYMI": "M",
    "RUDDI MELY": "F",
    "SULI MERCEDES": "F",
    "TEDI RILDO": "M",
    "SULBE": "M",
    "ALLENDE GUIMARAY": "M",
    "ANDRADE": "M",
    "ANDRADE GODOFREDO": "M",
    "ANDRADE OLIGARIO": "M",
    "ARCE BAYLON": "M",
    "ARCE CARLOS": "M",
    "ARCE ERMINES": "M",
    "ARCE NICIAL": "M",
    "BELARDE": "M",
    "BEQUE": "M",
    "DONAIRE DOMINGUEZ": "M",
    "MANRIQUE ROELIS": "M",
    "MANRIQUE ULDARICO": "M",
    "PONCE GILBERTO": "M",
    # --- Remaining 19 names from first run ---
    "AYDU": "F",           # variant of Ayda
    "CARLE\u00d1I": "F",   # Carleñi - female
    "GERALDINI YASMIN": "F",  # Yasmin is female
    "GERCE ROBLES": "M",     # male context
    "GO\u00d1E": "M",        # Goñe - male
    "HANANI": "F",            # female in Peruvian context
    "LEEMKE": "F",            # female variant
    "LEYDU": "F",             # variant of Leydi
    "LINEE DANY": "M",        # Dany male context
    "LISSE MARGOT": "F",      # Margot is female
    "MEQUI": "M",             # male
    "MILE\u00d1E": "F",       # Mileñe - female
    "NESSI-YU": "F",          # female
    "PERU LUZMERIDA": "F",    # Luzmerida is female
    "YALU": "F",              # female
    "YANDERI YENAVIERE": "F", # female
    "YARME": "M",             # male
    "YODE": "M",              # male
    "YUSTE DANI": "M",        # male
    "\u00d1URI MARLENE": "F", # Ñuri Marlene - female
    "E\u00d1E PILAR": "F",   # Eñe Pilar - female
}

# Additional single-token ambiguous names resolved
SINGLE_TOKEN_OVERRIDES = {
    "ABADIE": "F",
    "ABATE": "M",
    "AMADORI": "M",
    "DEYVI": "M",
    "DUANI": "F",
    "ENOE": "F",
    "ERLI": "F",
    "GEIRI": "F",
    "GELI": "F",
    "GENI": "F",
    "GIOVANI": "M",
    "GLADI": "F",
    "JERLI": "F",
    "JESI": "F",
    "LEBI": "F",
    "LENNI": "F",
    "LILE": "F",
    "LLINE": "F",
    "LOYDE": "F",
    "LUANI": "F",
    "LUDI": "F",
    "MELI": "F",
    "MECHI": "F",
    "MONLLI": "F",
    "NARDI": "F",
    "NOIMI": "F",
    "PARRALI": "F",
    "POLI": "M",
    "RONE": "M",
    "ROMNI": "M",
    "RORI": "M",
    "ROSSI": "F",
    "RULI": "M",
    "SOLI": "F",
    "SORI": "F",
    "TESI": "F",
    "TATI": "F",
    "YEMI": "F",
    "YENI": "F",
    "YENNI": "F",
    "YOLI": "F",
    "YONE": "M",
    "YONNI": "M",
    "YOVANNI": "F",
    "YRENE": "F",
    "DIONE": "M",
    "DIONI": "M",
    "JOYSI": "F",
    "NUMERABLE": "M",
    "MINAU": "M",
    "BERCI": "F",
    "ISOE": "M",
    "INOE": "F",
    "RODE": "F",
    "ROSALI": "F",
    "ROSMERI": "F",
    "LLERME": "F",
    "LLENI": "F",
    "LLERI": "F",
    "LLIMI": "F",
    "LLOCILU": "M",
    "LLOHANI": "F",
    "LLOVANI": "F",
    "HEYDI": "F",
    "LEIDI": "F",
    "LEYDI": "F",
    "LEYSI": "F",
    "DAYSI": "F",
    "DEYSI": "F",
    "DEYSSI": "F",
    "DEYCI": "F",
    "DAISI": "F",
    "DEISI": "F",
    "DAYCI": "F",
    "MARLENI": "F",
    "DENISSE": "F",
    "DENNISE": "F",
    "DENNISSE": "F",
    "JACKIE": "F",
    "JAKELINE": "F",
    "YACKELINE": "F",
    "YAQUELINE": "F",
    "LESLYE": "F",
    "JAQUELYNE": "F",
    "JHEYMI": "F",
    "JHOYSI": "F",
    "KATERINE": "F",
    "KATHERINNE": "F",
    "CORALI": "F",
    "NEYSI": "F",
    "NEYVI": "F",
    "NEYLI": "F",
    "NEYDI": "F",
    "NEYCI": "F",
    "NEYDE": "F",
    "NAHOMI": "F",
    "NELSI": "F",
    "NELVI": "F",
    "MADALEYNE": "F",
    "MADELEYNE": "F",
    "MEDALI": "F",
    "MILENE": "F",
    "MILENI": "F",
    "CLEDI": "F",
    "CLEIDI": "F",
    "CLEYDI": "F",
    "CLENI": "F",
    "DELMI": "F",
    "DELSI": "F",
    "DELCI": "F",
    "DAYSE": "F",
    "DEYSE": "F",
    "DEYNI": "F",
    "DEYMI": "F",
    "DEYLI": "F",
    "ESTEFANI": "F",
    "ESTEFANIE": "F",
    "ESTEFFANI": "F",
    "ESTHEFANI": "F",
    "STEPHANI": "F",
    "STHEFANI": "F",
    "ESTEPHANI": "F",
    "GREYSI": "F",
    "GREYSSI": "F",
    "GLEYDI": "F",
    "GLENI": "F",
    "KEYLI": "F",
    "KELI": "F",
    "ANALI": "F",
    "YAMELI": "F",
    "YAMILI": "F",
    "YERALDINE": "F",
    "YOSSELINE": "F",
    "JHAQUELINE": "F",
    "VERENICE": "F",
    "BRIGGITTE": "F",
    "DEISSI": "F",
    "DANAE": "F",
    "DANITZE": "F",
    "DARLI": "F",
    "FRIDE": "F",
    "DAGNE": "F",
    "DAGNI": "F",
    "DILCE": "F",
    "DILME": "F",
    "DORALE": "F",
    "DORALI": "F",
    "DORSI": "F",
    "JHUDI": "F",
    "COTI": "F",
    "BERCI": "F",
    "YUDI": "F",
    "CLIDE": "F",
    "MARLE": "F",
    "MERCI": "F",
    "NEPSI": "F",
    "YOLE": "F",
    "YOSELI": "F",
    "PEGGI": "F",
    "PEGUI": "F",
    "PRELI": "F",
    "QUELI": "F",
    "SARAHI": "F",
    "SELENNE": "F",
    "SILENI": "F",
    "STEISI": "F",
    "STEYSI": "F",
    "VASTI": "F",
    "XIOMI": "F",
    "ZURISADAI": "F",
    "LIZZETTE": "F",
    "LIZZETE": "F",
    "NOYMI": "F",
    "NANSI": "F",
    "IRENI": "F",
    "IBONI": "F",
    "HAIDEE": "F",
    "LIANI": "F",
    "LIDANI": "F",
    "LIDI": "F",
    "LIVIDI": "F",
    "LUZDARI": "F",
    "LUZMERI": "F",
    "MARYLU": "F",
    "MARDELI": "F",
    "MARISELI": "F",
    "MARILI": "F",
    "MELANNIE": "F",
    "MELVI": "F",
    "MIRELI": "F",
    "MIRLENE": "F",
    "ROSMELI": "F",
    "ROSANI": "F",
    "RUSMERI": "F",
    "SELENI": "F",
    "SHARMELI": "F",
    "YUVI": "F",
    "YUNELI": "F",
    "ZELENI": "F",
    "ZULI": "F",
    "IVETE": "F",
    "YVONE": "F",
    "FEBE": "F",
    "FEBI": "F",
    "JOYCI": "F",
    "JEANNI": "F",
    "JOAMI": "F",
    "MAGHI": "F",
    "MAHUMI": "F",
    "MOYSI": "F",
    "OVELI": "F",
    "DEYSERLI": "F",
    "DERMALI": "F",
    "GERSI": "F",
    "GESSI": "F",
    "LEAMERSI": "F",
    "LACENI": "F",
    "MAPI": "F",
    "MARFI": "F",
    "MARIOLI": "F",
    "MARSHELLI": "F",
    "MAYORI": "F",
    "MARUVENI": "F",
    "MAYYOLI": "F",
    "ROSSLI": "F",
    "RUTMELI": "F",
    "RUTHALI": "F",
    "SARMELI": "F",
    "SUGELI": "F",
    "SUGEILI": "F",
    "SUSELE": "F",
    "SUSETI": "F",
    "YABELI": "F",
    "YANALI": "F",
    "YANELE": "F",
    "YANELI": "F",
    "YARSELI": "F",
    "YASERI": "F",
    "YEIVI": "F",
    "YELE": "F",
    "YELI": "F",
    "YERENI": "F",
    "YESSI": "F",
    "YOCELI": "F",
    "YOLENI": "F",
    "ZIMI": "F",
    "ZOCI": "F",
    "ZOIDELI": "F",
    "ZOMELI": "F",
    "ZUELI": "F",
}


def classify_gender(name):
    """Classify gender of a name using token dictionaries and heuristics.

    Returns 'M', 'F', or '' if truly ambiguous.
    """
    if not isinstance(name, str) or not name.strip():
        return ""

    name = name.strip().upper()

    # 1. Check full-name override map first
    if name in OVERRIDE_MAP:
        return OVERRIDE_MAP[name]

    # 2. Check single-token overrides
    tokens = name.split()
    first = tokens[0]

    if name in SINGLE_TOKEN_OVERRIDES:
        return SINGLE_TOKEN_OVERRIDES[name]
    if first in SINGLE_TOKEN_OVERRIDES:
        return SINGLE_TOKEN_OVERRIDES[first]

    # 3. Check first token against dictionaries
    if first in FEMALE_TOKENS:
        return "F"
    if first in MALE_TOKENS:
        return "M"

    # 4. Check subsequent tokens (second name often clarifies)
    for token in tokens[1:]:
        if token in {"DE", "LA", "DEL", "LOS", "LAS", "VDA", "VDA.DE"}:
            continue
        if token in FEMALE_TOKENS:
            return "F"
        if token in MALE_TOKENS:
            return "M"

    # 5. Heuristic: common Peruvian name endings
    # Names ending in common female suffixes
    if first.endswith(("ILDA", "ILDE", "ILDA", "ELDE", "ILDE")):
        return "F"
    if first.endswith(("INE", "INE", "YNNE", "INNE", "ENNE")):
        return "F"

    return ""


# =====================================================================
# EXCEL FORMATTING (identical to crear_excel_alcaldes.py)
# =====================================================================
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
EVEN_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ODD_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
CENTER_COLS = {"ubigeo", "sexo", "total_votos", "dni"}


def format_excel(xlsx_path, year):
    """Apply formatting to an Excel file."""
    wb = load_workbook(xlsx_path)
    ws = wb.active

    num_cols = ws.max_column
    num_rows = ws.max_row
    headers = [ws.cell(row=1, column=c).value for c in range(1, num_cols + 1)]

    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    for row_idx in range(2, num_rows + 1):
        fill = EVEN_FILL if row_idx % 2 == 0 else ODD_FILL
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            cell.fill = fill
            col_name = headers[col_idx - 1]
            if col_name and col_name.lower() in CENTER_COLS:
                cell.alignment = CENTER_ALIGN
            else:
                cell.alignment = LEFT_ALIGN

    for col_idx in range(1, num_cols + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, min(num_rows + 1, 1000)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        adjusted = min(max_len + 3, 45)
        ws.column_dimensions[col_letter].width = max(adjusted, 8)

    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows}"
    ws.freeze_panes = "A2"
    wb.save(xlsx_path)


# =====================================================================
# MAIN
# =====================================================================
def main():
    years_to_fix = [2006, 2010, 2014]
    all_years = [2006, 2010, 2014, 2018, 2022]

    # Build name->gender cache from all missing names
    print("=" * 70)
    print("completar_sexo.py - Completar genero faltante")
    print("=" * 70)

    total_completed = 0
    total_remaining = 0
    remaining_names = set()

    # --- Step 1: Update alcaldes_con_votos CSVs ---
    print("\n-- Paso 1: Actualizar alcaldes_con_votos --")
    for year in years_to_fix:
        csv_path = os.path.join(ALCALDES_DIR, f"alcaldes_con_votos_{year}.csv")
        df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
        df.columns = [c.strip() for c in df.columns]

        mask = df["sexo"].isna() | (df["sexo"].str.strip() == "")
        n_missing_before = mask.sum()

        if n_missing_before > 0:
            df.loc[mask, "sexo"] = df.loc[mask, "nombres"].apply(classify_gender)

        mask_after = df["sexo"].isna() | (df["sexo"].str.strip() == "")
        n_missing_after = mask_after.sum()
        completed = n_missing_before - n_missing_after

        total_completed += completed
        total_remaining += n_missing_after

        if n_missing_after > 0:
            remaining_names.update(
                df.loc[mask_after, "nombres"].dropna().str.strip().unique()
            )

        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"  {year}: {n_missing_before} vacios -> {completed} completados, {n_missing_after} restantes")

    # --- Step 2: Update candidatos_completos CSVs ---
    print("\n-- Paso 2: Actualizar candidatos_completos --")
    for year in years_to_fix:
        csv_path = os.path.join(CANDIDATOS_DIR, f"candidatos_{year}.csv")
        df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
        df.columns = [c.strip() for c in df.columns]

        mask = df["sexo"].isna() | (df["sexo"].str.strip() == "")
        n_missing_before = mask.sum()

        if n_missing_before > 0:
            df.loc[mask, "sexo"] = df.loc[mask, "nombres"].apply(classify_gender)

        mask_after = df["sexo"].isna() | (df["sexo"].str.strip() == "")
        n_missing_after = mask_after.sum()
        completed = n_missing_before - n_missing_after

        total_completed += completed
        total_remaining += n_missing_after

        if n_missing_after > 0:
            remaining_names.update(
                df.loc[mask_after, "nombres"].dropna().str.strip().unique()
            )

        df.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"  {year}: {n_missing_before} vacios -> {completed} completados, {n_missing_after} restantes")

    # --- Step 3: Regenerate all 5 Excel files ---
    print("\n-- Paso 3: Regenerar 5 archivos Excel --")
    for year in all_years:
        csv_path = os.path.join(ALCALDES_DIR, f"alcaldes_con_votos_{year}.csv")
        xlsx_path = os.path.join(BASE_DIR, f"alcaldes_{year}.xlsx")

        df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        df.to_excel(xlsx_path, index=False, sheet_name=f"Alcaldes {year}", engine="openpyxl")
        format_excel(xlsx_path, year)
        print(f"  {year}: {xlsx_path} ({len(df)} filas)")

    # --- Step 4: Summary ---
    print("\n" + "=" * 70)
    print("RESUMEN")
    print("=" * 70)
    print(f"  Total completados (alcaldes + candidatos): {total_completed}")
    print(f"  Total restantes sin clasificar:            {total_remaining}")

    if remaining_names:
        print(f"\n  Nombres unicos sin clasificar ({len(remaining_names)}):")
        for n in sorted(remaining_names):
            print(f"    {n}")

    # Verification: show final sexo stats per year
    print("\n-- Estadisticas finales de sexo por archivo --")
    for year in all_years:
        csv_path = os.path.join(ALCALDES_DIR, f"alcaldes_con_votos_{year}.csv")
        df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        n = len(df)
        n_m = (df["sexo"] == "M").sum()
        n_f = (df["sexo"] == "F").sum()
        n_empty = n - n_m - n_f
        pct = (n_m + n_f) / n * 100 if n > 0 else 0
        print(f"  Alcaldes {year}: {n} total | M={n_m} F={n_f} vacios={n_empty} | {pct:.1f}% completo")

    for year in years_to_fix:
        csv_path = os.path.join(CANDIDATOS_DIR, f"candidatos_{year}.csv")
        df = pd.read_csv(csv_path, encoding="utf-8-sig", dtype=str)
        df.columns = [c.strip() for c in df.columns]
        n = len(df)
        n_m = (df["sexo"] == "M").sum()
        n_f = (df["sexo"] == "F").sum()
        n_empty = n - n_m - n_f
        pct = (n_m + n_f) / n * 100 if n > 0 else 0
        print(f"  Candidatos {year}: {n} total | M={n_m} F={n_f} vacios={n_empty} | {pct:.1f}% completo")

    print("\nListo!")


if __name__ == "__main__":
    main()
